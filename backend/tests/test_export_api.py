from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from app.models.employee import Employee
from app.api.export import _build_employee_report_directory
from app.api.importer import _build_employee_report_directory as _build_import_employee_report_directory


def test_export_timesheet_excel_success(client, seed_timesheet_data):
    period_start = seed_timesheet_data["period_start"].isoformat()
    period_end = seed_timesheet_data["period_end"].isoformat()

    response = client.get("/api/export/timesheet", params={"period_start": period_start, "period_end": period_end})

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment; filename=\"timesheet_2026-04-23_2026-05-22.xlsx\"" in response.headers.get(
        "content-disposition", ""
    )
    assert response.content[:2] == b"PK"

    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook["Timesheet"]
    assert sheet["G6"].value == "EMPLOYEE TIMESHEET FROM 23/04/2026 TO 22/05/2026"
    assert sheet["A7"].value == "STT"
    assert sheet["B7"].value == "Họ Và Tên"
    assert sheet["C7"].value == 23
    assert sheet["E7"].value == "25\nT7"
    assert sheet["F7"].value == "26\nCN"
    assert sheet["BO7"].value == "Nghỉ\nkhông\nlương"
    assert sheet["BP7"].value == "Nghỉ\nhưởng\nlương"
    assert sheet["BQ7"].value == "Ngày phép\ncòn lại\ntháng trước"
    assert sheet["BR7"].value == "Ngày phép\ntháng này"
    assert sheet["BS7"].value == "Ngày phép\ncòn lại"
    assert sheet["A8"].value == "E001"
    assert sheet["B8"].value == "Nguyen Van A"
    assert sheet["BO8"].value == 0
    assert sheet["BP8"].value == 1
    assert sheet["BQ8"].value == 0
    assert sheet["BR8"].value == 1
    assert sheet["BS8"].value == 0
    assert sheet["F8"].fill.fgColor.rgb in {"00E2F0D9", "FFE2F0D9"}
    assert sheet["C8"].border.left.style == "thin"


def test_export_timesheet_uses_vietnamese_employee_name(client, seed_timesheet_data, seed_basic_employees, db_session):
    worker = seed_basic_employees["worker"]
    worker.full_name = "ĐẶNG HOÀI BẢO"
    worker.notion_name = "Baron"
    db_session.commit()

    response = client.get(
        "/api/export/timesheet",
        params={
            "period_start": seed_timesheet_data["period_start"].isoformat(),
            "period_end": seed_timesheet_data["period_end"].isoformat(),
        },
    )

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    assert workbook["Timesheet"]["B8"].value == "ĐẶNG HOÀI BẢO"


def test_uploaded_report_name_mapping_prefers_vietnamese_employee_name(db_session):
    employee = Employee(
        machine_employee_id="38",
        full_name="ĐẶNG HOÀI BẢO",
        notion_name="Baron",
        annual_leave_quota=12,
        annual_leave_used=0,
        paid_leave_balance=0,
        unpaid_leave_balance=0,
        is_active=True,
    )
    db_session.add(employee)
    db_session.commit()

    directory = _build_employee_report_directory(db_session)
    assert directory["38"]["full_name"] == "ĐẶNG HOÀI BẢO"


def test_import_preview_name_mapping_prefers_vietnamese_employee_name(db_session):
    employee = Employee(
        machine_employee_id="26",
        full_name="Nguyễn Thanh Đạt",
        notion_name="TOMMY DAT",
        annual_leave_quota=12,
        annual_leave_used=0,
        paid_leave_balance=0,
        unpaid_leave_balance=0,
        is_active=True,
    )
    db_session.add(employee)
    db_session.commit()

    directory = _build_import_employee_report_directory(db_session)
    assert directory["26"]["full_name"] == "Nguyễn Thanh Đạt"


def test_export_attendance_json_report_from_uploaded_workbook(client):
    cycle_days = list(range(23, 32)) + list(range(1, 23))
    schedule_df = pd.DataFrame(
        [
            ["ID", "Ten", "Phong ban", *cycle_days],
            ["2", "NGUYEN THANH TR", "Not Set1", *[1 if day in {23, 24, 26, 27} else "" for day in cycle_days]],
        ]
    )
    profile_df = pd.DataFrame(
        [
            ["ID", "Ten", "Phong ban", *cycle_days],
            ["2", "NGUYEN THANH TR", "Not Set1", *["" for _ in cycle_days]],
            ["", "", "", *["08:15\n17:30" if day == 23 else "08:45*\n17:30" if day == 24 else "08:15" if day == 26 else "" for day in cycle_days]],
        ]
    )
    abnormal_df = pd.DataFrame(
        [
            ["ID", "Ngay", "Buoi 1 Vao lam", "Buoi 1 Ra nghi", "Thoi gian tre", "Thoi gian som"],
            ["2", "2026-03-24", "08:45", "17:30", 15, 0],
            ["2", "2026-03-26", "08:15", "Bỏ lỡ", 0, 0],
            ["2", "2026-03-27", "Bỏ lỡ", "Bỏ lỡ", 0, 0],
        ]
    )
    summary_df = pd.DataFrame(
        [
            ["ID", "Ten", "Phong ban", "Tong so phut di muon trong thang", "Tong so ngay vang mat"],
            ["2", "NGUYEN THANH TR", "Not Set1", 15, 1],
        ]
    )
    checkin_report_df = pd.DataFrame(
        [
            ["ID", "Ten", "Phong ban", "Ngay", "Gio vao", "Gio ra"],
            ["2", "NGUYEN THANH TR", "Not Set1", "2026-03-23", "08:15", "17:30"],
        ]
    )

    stream = BytesIO()
    with pd.ExcelWriter(stream, engine="openpyxl") as writer:
        schedule_df.to_excel(writer, index=False, header=False, sheet_name="Bảng thông tin lịch trình")
        checkin_report_df.to_excel(writer, index=False, header=False, sheet_name="Báo cáo check-in")
        abnormal_df.to_excel(writer, index=False, header=False, sheet_name="Báo cáo bất thường")
        profile_df.to_excel(writer, index=False, header=False, sheet_name="Hồ sơ check-in")
        summary_df.to_excel(writer, index=False, header=False, sheet_name="Bảng tóm tắt check-in")
    stream.seek(0)

    response = client.post(
        "/api/export/attendance-json-report",
        files={
            "file": (
                "attendance_full_cycle.xlsx",
                stream,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"period_start": "2026-03-23"},
    )

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook["Timesheet"]
    assert sheet["G6"].value == "EMPLOYEE TIMESHEET FROM 23/03/2026 TO 22/04/2026"
    assert sheet["B8"].value == "NGUYEN THANH TR"
    assert sheet["C8"].value == 1.0
    assert sheet["D8"].value == 1.0
    assert sheet["F8"].value == 1.0
    assert sheet["G8"].value == 0.0
    assert sheet["AL8"].value == "X"
    assert sheet["AM8"].value == "X"
    assert sheet["AO8"].value == "X"
    assert sheet["AP8"].value == "Ro"


def test_export_attendance_json_report_skips_schedule_only_employees(client):
    cycle_days = list(range(23, 32)) + list(range(1, 23))
    schedule_df = pd.DataFrame(
        [
            ["ID", "Ten", "Phong ban", *cycle_days],
            ["2", "NGUYEN THANH TR", "Not Set1", *[1 if day in {23, 24} else "" for day in cycle_days]],
            ["9", "LE VAN MOCK", "Kho Mock", *[1 if day in {23, 24} else "" for day in cycle_days]],
        ]
    )
    profile_df = pd.DataFrame(
        [
            ["ID", "Ten", "Phong ban", *cycle_days],
            ["2", "NGUYEN THANH TR", "Not Set1", *["" for _ in cycle_days]],
            ["", "", "", *["08:15\n17:30" if day == 23 else "08:45\n17:30" if day == 24 else "" for day in cycle_days]],
        ]
    )
    abnormal_df = pd.DataFrame(
        [
            ["ID", "Ngay", "Buoi 1 Vao lam", "Buoi 1 Ra nghi", "Thoi gian tre", "Thoi gian som"],
            ["2", "2026-03-24", "08:45", "17:30", 15, 0],
        ]
    )
    summary_df = pd.DataFrame(
        [
            ["ID", "Ten", "Phong ban", "Tong so phut di muon trong thang", "Tong so ngay vang mat"],
            ["2", "NGUYEN THANH TR", "Not Set1", 15, 0],
            ["9", "LE VAN MOCK", "Kho Mock", 0, 0],
        ]
    )
    checkin_report_df = pd.DataFrame(
        [
            ["ID", "Ten", "Phong ban", "Ngay", "Gio vao", "Gio ra"],
            ["2", "NGUYEN THANH TR", "Not Set1", "2026-03-23", "08:15", "17:30"],
        ]
    )

    stream = BytesIO()
    with pd.ExcelWriter(stream, engine="openpyxl") as writer:
        schedule_df.to_excel(writer, index=False, header=False, sheet_name="Bảng thông tin lịch trình")
        checkin_report_df.to_excel(writer, index=False, header=False, sheet_name="Báo cáo check-in")
        abnormal_df.to_excel(writer, index=False, header=False, sheet_name="Báo cáo bất thường")
        profile_df.to_excel(writer, index=False, header=False, sheet_name="Hồ sơ check-in")
        summary_df.to_excel(writer, index=False, header=False, sheet_name="Bảng tóm tắt check-in")
    stream.seek(0)

    response = client.post(
        "/api/export/attendance-json-report",
        files={
            "file": (
                "attendance_schedule_only_employee.xlsx",
                stream,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"period_start": "2026-03-23"},
    )

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook["Timesheet"]
    assert sheet["B8"].value == "NGUYEN THANH TR"
    assert sheet["B9"].value is None


def test_export_attendance_json_report_applies_approved_notion_submission(client):
    cycle_days = list(range(23, 32)) + list(range(1, 23))
    schedule_df = pd.DataFrame(
        [
            ["ID", "Ten", "Phong ban", *cycle_days],
            ["2", "NGUYEN THANH TR", "Not Set1", *[1 if day in {23, 24, 26, 27} else "" for day in cycle_days]],
        ]
    )
    profile_df = pd.DataFrame(
        [
            ["ID", "Ten", "Phong ban", *cycle_days],
            ["2", "NGUYEN THANH TR", "Not Set1", *["" for _ in cycle_days]],
            ["", "", "", *["08:15\n17:30" if day == 23 else "08:45*\n17:30" if day == 24 else "08:15" if day == 26 else "" for day in cycle_days]],
        ]
    )
    abnormal_df = pd.DataFrame(
        [
            ["ID", "Ngay", "Buoi 1 Vao lam", "Buoi 1 Ra nghi", "Thoi gian tre", "Thoi gian som"],
            ["2", "2026-03-24", "08:45", "17:30", 15, 0],
            ["2", "2026-03-26", "08:15", "Bỏ lỡ", 0, 0],
            ["2", "2026-03-27", "Bỏ lỡ", "Bỏ lỡ", 0, 0],
        ]
    )
    summary_df = pd.DataFrame(
        [
            ["ID", "Ten", "Phong ban", "Tong so phut di muon trong thang", "Tong so ngay vang mat"],
            ["2", "NGUYEN THANH TR", "Not Set1", 15, 1],
        ]
    )
    checkin_report_df = pd.DataFrame(
        [
            ["ID", "Ten", "Phong ban", "Ngay", "Gio vao", "Gio ra"],
            ["2", "NGUYEN THANH TR", "Not Set1", "2026-03-23", "08:15", "17:30"],
        ]
    )

    workbook_stream = BytesIO()
    with pd.ExcelWriter(workbook_stream, engine="openpyxl") as writer:
        schedule_df.to_excel(writer, index=False, header=False, sheet_name="Bảng thông tin lịch trình")
        checkin_report_df.to_excel(writer, index=False, header=False, sheet_name="Báo cáo check-in")
        abnormal_df.to_excel(writer, index=False, header=False, sheet_name="Báo cáo bất thường")
        profile_df.to_excel(writer, index=False, header=False, sheet_name="Hồ sơ check-in")
        summary_df.to_excel(writer, index=False, header=False, sheet_name="Bảng tóm tắt check-in")
    workbook_stream.seek(0)

    notion_csv = "\n".join(
        [
            "Name,Tên nhân viên,Leave Balance,Lý do Nghỉ,Thời Gian,Số Ngày Nghỉ,Trạng Thái",
            "Leave Request,DOCS - NGUYEN THANH TR,DOCS - NGUYEN THANH TR,Cá nhân,03/27/2026 8:00 AM (GMT+7) → 5:30 PM,1,Approved",
        ]
    ).encode("utf-8")

    response = client.post(
        "/api/export/attendance-json-report",
        files={
            "file": (
                "attendance_full_cycle.xlsx",
                workbook_stream,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            "notion_file": ("leave_request.csv", notion_csv, "text/csv"),
        },
        data={"period_start": "2026-03-23"},
    )

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook["Timesheet"]
    assert sheet["AP8"].value == "P"
    assert sheet["BQ8"].value == 0
    assert sheet["BR8"].value == 1


def test_export_attendance_json_report_uses_employee_notion_mapping_from_database(client, db_session):
    employee = Employee(
        machine_employee_id="2",
        full_name="DOCS - PARADO QUANG",
        notion_name="DOCS - PARADO QUANG",
        department_code="OPS",
        department_name="Operations",
        annual_leave_quota=12,
        annual_leave_used=0,
        paid_leave_balance=0,
        unpaid_leave_balance=0,
        is_active=True,
    )
    db_session.add(employee)
    db_session.commit()

    cycle_days = list(range(23, 32)) + list(range(1, 23))
    schedule_df = pd.DataFrame(
        [
            ["ID", "Ten", "Phong ban", *cycle_days],
            ["2", "MAY CHAM CONG 902", "Operations", *[1 if day in {23, 24, 26, 27} else "" for day in cycle_days]],
        ]
    )
    profile_df = pd.DataFrame(
        [
            ["ID", "Ten", "Phong ban", *cycle_days],
            ["2", "MAY CHAM CONG 902", "Operations", *["" for _ in cycle_days]],
            ["", "", "", *["08:15\n17:30" if day == 23 else "08:45*\n17:30" if day == 24 else "08:15" if day == 26 else "" for day in cycle_days]],
        ]
    )
    abnormal_df = pd.DataFrame(
        [
            ["ID", "Ngay", "Buoi 1 Vao lam", "Buoi 1 Ra nghi", "Thoi gian tre", "Thoi gian som"],
            ["2", "2026-03-24", "08:45", "17:30", 15, 0],
            ["2", "2026-03-26", "08:15", "Bỏ lỡ", 0, 0],
            ["2", "2026-03-27", "Bỏ lỡ", "Bỏ lỡ", 0, 0],
        ]
    )
    summary_df = pd.DataFrame(
        [
            ["ID", "Ten", "Phong ban", "Tong so phut di muon trong thang", "Tong so ngay vang mat"],
            ["2", "MAY CHAM CONG 902", "Operations", 15, 1],
        ]
    )
    checkin_report_df = pd.DataFrame(
        [
            ["ID", "Ten", "Phong ban", "Ngay", "Gio vao", "Gio ra"],
            ["2", "MAY CHAM CONG 902", "Operations", "2026-03-23", "08:15", "17:30"],
        ]
    )

    workbook_stream = BytesIO()
    with pd.ExcelWriter(workbook_stream, engine="openpyxl") as writer:
        schedule_df.to_excel(writer, index=False, header=False, sheet_name="Bảng thông tin lịch trình")
        checkin_report_df.to_excel(writer, index=False, header=False, sheet_name="Báo cáo check-in")
        abnormal_df.to_excel(writer, index=False, header=False, sheet_name="Báo cáo bất thường")
        profile_df.to_excel(writer, index=False, header=False, sheet_name="Hồ sơ check-in")
        summary_df.to_excel(writer, index=False, header=False, sheet_name="Bảng tóm tắt check-in")
    workbook_stream.seek(0)

    notion_csv = "\n".join(
        [
            "Name,Tên nhân viên,Leave Balance,Lý do Nghỉ,Thời Gian,Số Ngày Nghỉ,Trạng Thái",
            "Leave Request,DOCS - PARADO QUANG,DOCS - PARADO QUANG,Cá nhân,03/27/2026 8:00 AM (GMT+7) → 5:30 PM,1,Approved",
        ]
    ).encode("utf-8")

    response = client.post(
        "/api/export/attendance-json-report",
        files={
            "file": (
                "attendance_full_cycle.xlsx",
                workbook_stream,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            "notion_file": ("leave_request.csv", notion_csv, "text/csv"),
        },
        data={"period_start": "2026-03-23"},
    )

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook["Timesheet"]
    assert sheet["B8"].value == "DOCS - PARADO QUANG"
    assert sheet["AP8"].value == "P"


def test_export_attendance_json_report_uses_employee_full_name_from_database_without_notion_file(client, db_session):
    employee = Employee(
        machine_employee_id="2",
        full_name="ERIC QUAN",
        notion_name="ERIC QUAN",
        department_code="OPS",
        department_name="Operations",
        annual_leave_quota=12,
        annual_leave_used=0,
        paid_leave_balance=0,
        unpaid_leave_balance=0,
        is_active=True,
    )
    db_session.add(employee)
    db_session.commit()

    cycle_days = list(range(23, 32)) + list(range(1, 23))
    schedule_df = pd.DataFrame(
        [
            ["ID", "Ten", "Phong ban", *cycle_days],
            ["2", "MAY CHAM CONG 902", "Operations", *[1 if day in {23, 24} else "" for day in cycle_days]],
        ]
    )
    profile_df = pd.DataFrame(
        [
            ["ID", "Ten", "Phong ban", *cycle_days],
            ["2", "MAY CHAM CONG 902", "Operations", *["" for _ in cycle_days]],
            ["", "", "", *["08:15\n17:30" if day == 23 else "08:45\n17:30" if day == 24 else "" for day in cycle_days]],
        ]
    )
    abnormal_df = pd.DataFrame(
        [
            ["ID", "Ngay", "Buoi 1 Vao lam", "Buoi 1 Ra nghi", "Thoi gian tre", "Thoi gian som"],
            ["2", "2026-03-24", "08:45", "17:30", 15, 0],
        ]
    )
    summary_df = pd.DataFrame(
        [
            ["ID", "Ten", "Phong ban", "Tong so phut di muon trong thang", "Tong so ngay vang mat"],
            ["2", "MAY CHAM CONG 902", "Operations", 15, 0],
        ]
    )
    checkin_report_df = pd.DataFrame(
        [
            ["ID", "Ten", "Phong ban", "Ngay", "Gio vao", "Gio ra"],
            ["2", "MAY CHAM CONG 902", "Operations", "2026-03-23", "08:15", "17:30"],
        ]
    )

    workbook_stream = BytesIO()
    with pd.ExcelWriter(workbook_stream, engine="openpyxl") as writer:
        schedule_df.to_excel(writer, index=False, header=False, sheet_name="Bảng thông tin lịch trình")
        checkin_report_df.to_excel(writer, index=False, header=False, sheet_name="Báo cáo check-in")
        abnormal_df.to_excel(writer, index=False, header=False, sheet_name="Báo cáo bất thường")
        profile_df.to_excel(writer, index=False, header=False, sheet_name="Hồ sơ check-in")
        summary_df.to_excel(writer, index=False, header=False, sheet_name="Bảng tóm tắt check-in")
    workbook_stream.seek(0)

    response = client.post(
        "/api/export/attendance-json-report",
        files={
            "file": (
                "attendance_full_cycle.xlsx",
                workbook_stream,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"period_start": "2026-03-23"},
    )

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook["Timesheet"]
    assert sheet["B8"].value == "ERIC QUAN"


def test_export_attendance_json_report_marks_half_day_leave_symbols_from_notion_time_ranges(client):
    cycle_days = list(range(23, 32)) + list(range(1, 23))
    schedule_df = pd.DataFrame(
        [
            ["ID", "Ten", "Phong ban", *cycle_days],
            ["2", "NGUYEN THANH TR", "Not Set1", *[1 if day in {26, 27, 31, 1} else "" for day in cycle_days]],
        ]
    )
    profile_df = pd.DataFrame(
        [
            ["ID", "Ten", "Phong ban", *cycle_days],
            ["2", "NGUYEN THANH TR", "Not Set1", *["" for _ in cycle_days]],
            [
                "",
                "",
                "",
                *[
                    "08:15\n12:00" if day == 26 else "13:05\n17:30" if day == 27 else "" for day in cycle_days
                ],
            ],
        ]
    )
    abnormal_df = pd.DataFrame(
        [
            ["ID", "Ngay", "Buoi 1 Vao lam", "Buoi 1 Ra nghi", "Thoi gian tre", "Thoi gian som"],
        ]
    )
    summary_df = pd.DataFrame(
        [
            ["ID", "Ten", "Phong ban", "Tong so phut di muon trong thang", "Tong so ngay vang mat"],
            ["2", "NGUYEN THANH TR", "Not Set1", 0, 0],
        ]
    )
    checkin_report_df = pd.DataFrame(
        [
            ["ID", "Ten", "Phong ban", "Ngay", "Gio vao", "Gio ra"],
            ["2", "NGUYEN THANH TR", "Not Set1", "2026-03-26", "08:15", "12:00"],
            ["2", "NGUYEN THANH TR", "Not Set1", "2026-03-27", "13:05", "17:30"],
        ]
    )

    workbook_stream = BytesIO()
    with pd.ExcelWriter(workbook_stream, engine="openpyxl") as writer:
        schedule_df.to_excel(writer, index=False, header=False, sheet_name="Bảng thông tin lịch trình")
        checkin_report_df.to_excel(writer, index=False, header=False, sheet_name="Báo cáo check-in")
        abnormal_df.to_excel(writer, index=False, header=False, sheet_name="Báo cáo bất thường")
        profile_df.to_excel(writer, index=False, header=False, sheet_name="Hồ sơ check-in")
        summary_df.to_excel(writer, index=False, header=False, sheet_name="Bảng tóm tắt check-in")
    workbook_stream.seek(0)

    notion_csv = "\n".join(
        [
            "Name,Tên nhân viên,Leave Balance,Lý do Nghỉ,Thời Gian,Số Ngày Nghỉ,Trạng Thái",
            "Leave Request,DOCS - NGUYEN THANH TR,DOCS - NGUYEN THANH TR,Cá nhân,03/26/2026 12:00 PM (GMT+7) → 5:30 PM,0.5,Approved",
            "Leave Request,DOCS - NGUYEN THANH TR,DOCS - NGUYEN THANH TR,Cá nhân,03/27/2026 8:00 AM (GMT+7) → 1:00 PM,0.5,Approved",
            "Leave Request,DOCS - NGUYEN THANH TR,DOCS - NGUYEN THANH TR,Cá nhân,03/31/2026 8:30 AM (GMT+7) → 12:00 PM,0.5,Approved",
            "Leave Request,DOCS - NGUYEN THANH TR,DOCS - NGUYEN THANH TR,Cá nhân,04/01/2026 12:00 PM (GMT+7) → 5:30 PM,0.5,Approved",
        ]
    ).encode("utf-8")

    response = client.post(
        "/api/export/attendance-json-report",
        files={
            "file": (
                "attendance_half_day.xlsx",
                workbook_stream,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            "notion_file": ("leave_request.csv", notion_csv, "text/csv"),
        },
        data={"period_start": "2026-03-23"},
    )

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook["Timesheet"]
    assert sheet["AO8"].value == "X/P"
    assert sheet["AP8"].value == "P/X"
    assert sheet["AT8"].value == "P/Ro"
    assert sheet["AU8"].value == "Ro/P"


def test_export_attendance_json_report_keeps_full_day_symbol_when_leave_units_are_1_day(client):
    cycle_days = list(range(23, 32)) + list(range(1, 23))
    schedule_df = pd.DataFrame(
        [
            ["ID", "Ten", "Phong ban", *cycle_days],
            ["61", "kimkt", "Accounting", *[1 if day in {23, 13, 22} else "" for day in cycle_days]],
        ]
    )
    profile_df = pd.DataFrame(
        [
            ["ID", "Ten", "Phong ban", *cycle_days],
            ["61", "kimkt", "Accounting", *["" for _ in cycle_days]],
            ["", "", "", *["08:15\n17:30" if day == 23 else "" for day in cycle_days]],
        ]
    )
    abnormal_df = pd.DataFrame(
        [
            ["ID", "Ngay", "Buoi 1 Vao lam", "Buoi 1 Ra nghi", "Thoi gian tre", "Thoi gian som"],
        ]
    )
    summary_df = pd.DataFrame(
        [
            ["ID", "Ten", "Phong ban", "Tong so phut di muon trong thang", "Tong so ngay vang mat"],
            ["61", "kimkt", "Accounting", 0, 0],
        ]
    )
    checkin_report_df = pd.DataFrame(
        [
            ["ID", "Ten", "Phong ban", "Ngay", "Gio vao", "Gio ra"],
            ["61", "kimkt", "Accounting", "2026-03-23", "08:15", "17:30"],
        ]
    )

    workbook_stream = BytesIO()
    with pd.ExcelWriter(workbook_stream, engine="openpyxl") as writer:
        schedule_df.to_excel(writer, index=False, header=False, sheet_name="Bảng thông tin lịch trình")
        checkin_report_df.to_excel(writer, index=False, header=False, sheet_name="Báo cáo check-in")
        abnormal_df.to_excel(writer, index=False, header=False, sheet_name="Báo cáo bất thường")
        profile_df.to_excel(writer, index=False, header=False, sheet_name="Hồ sơ check-in")
        summary_df.to_excel(writer, index=False, header=False, sheet_name="Bảng tóm tắt check-in")
    workbook_stream.seek(0)

    notion_csv = "\n".join(
        [
            "Name,Tên nhân viên,Leave Balance,Lý do Nghỉ,Thời Gian,Số Ngày Nghỉ,Trạng Thái",
            "Leave Request,kimkt,kimkt,Đau ốm,04/13/2026 12:00 PM (GMT+7) → 5:30 PM,1.0,Approved",
            "Leave Request,kimkt,kimkt,Cá nhân,04/22/2026 8:00 AM (GMT+7) → 1:00 PM,1.0,Approved",
        ]
    ).encode("utf-8")

    response = client.post(
        "/api/export/attendance-json-report",
        files={
            "file": (
                "attendance_full_day_with_half_day_times.xlsx",
                workbook_stream,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            "notion_file": ("leave_request.csv", notion_csv, "text/csv"),
        },
        data={"period_start": "2026-03-23"},
    )

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook["Timesheet"]
    # 04/13/2026 12:00 PM -> 5:30 PM is afternoon leave, with no attendance it resolves to Ro/P
    assert sheet["BG8"].value == "Ro/P"
    # 04/22/2026 8:00 AM -> 1:00 PM is morning leave, with no attendance it resolves to P/Ro
    assert sheet["BP8"].value == "P/Ro"


def test_export_timesheet_excel_not_found(client):
    response = client.get(
        "/api/export/timesheet",
        params={"period_start": "2026-04-23", "period_end": "2026-05-22"},
    )
    assert response.status_code == 404
    assert response.json()["detail"] == "no timesheet data for selected period"
