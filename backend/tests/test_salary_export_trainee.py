from io import BytesIO

from openpyxl import load_workbook

from app.models.employee import Employee
from app.models.monthly_salary_input import MonthlySalaryInput
from app.services.salary import export_salary_report


def test_salary_export_lists_trainee_in_block_c_without_payroll_or_transfer(db_session):
    fulltime = Employee(
        machine_employee_id="PAY-A-1",
        full_name="Payroll Employee",
        employee_code="PAY-A-1",
        employee_type="FULLTIME",
        contract_salary=20_000_000,
        position="ACCOUNTING",
        account_number="111111",
        bank_name="TEST BANK",
        is_active=True,
        status="ACTIVE",
    )
    trainee = Employee(
        machine_employee_id="PAY-C-1",
        full_name="Trainee Roster Only",
        employee_code="PAY-C-1",
        employee_type="TRAINEE",
        contract_salary=50_000_000,
        position="TTS",
        account_number="222222",
        bank_name="TEST BANK",
        is_active=True,
        status="ACTIVE",
    )
    db_session.add_all([fulltime, trainee])
    db_session.flush()
    db_session.add_all(
        [
            MonthlySalaryInput(
                employee_id=fulltime.id,
                salary_period="2026-08",
                actual_working_days=23,
            ),
            MonthlySalaryInput(
                employee_id=trainee.id,
                salary_period="2026-08",
                actual_working_days=23,
                bonus=99_000_000,
                other_income=88_000_000,
            ),
        ]
    )
    db_session.commit()

    workbook = load_workbook(BytesIO(export_salary_report(db_session, "2026-08").getvalue()), data_only=False)
    salary_sheet = workbook["Employee salary"]
    bank_sheet = workbook["Bank Transfer"]
    sealink_sheet = workbook["Sealink Transfer"]

    block_c_row = next(
        row
        for row in range(1, salary_sheet.max_row + 1)
        if str(salary_sheet.cell(row=row, column=1).value or "").startswith("C. THỰC TẬP SINH")
    )
    trainee_row = next(
        row
        for row in range(block_c_row + 1, salary_sheet.max_row + 1)
        if salary_sheet.cell(row=row, column=3).value == "Trainee Roster Only"
    )
    total_row = next(
        row
        for row in range(trainee_row + 1, salary_sheet.max_row + 1)
        if salary_sheet.cell(row=row, column=1).value == "TỔNG CỘNG"
    )

    assert salary_sheet.cell(row=trainee_row, column=2).value == 23
    assert "không nhập số tiền" in salary_sheet.cell(row=trainee_row, column=6).value
    assert salary_sheet.cell(row=trainee_row, column=35).value is None
    assert salary_sheet.cell(row=total_row, column=35).value.startswith("=")

    bank_names = {
        bank_sheet.cell(row=row, column=2).value
        for row in range(1, bank_sheet.max_row + 1)
    }
    sealink_names = {
        sealink_sheet.cell(row=row, column=7).value
        for row in range(1, sealink_sheet.max_row + 1)
    }
    assert "Payroll Employee" in bank_names
    assert "Trainee Roster Only" not in bank_names
    assert "PAYROLL EMPLOYEE" in sealink_names
    assert "TRAINEE ROSTER ONLY" not in sealink_names
