from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from app.services.final_timesheet_report import (
    FinalTimesheetDailyInput,
    FinalTimesheetEmployeeInput,
    FinalTimesheetEntryInput,
    FinalTimesheetOffRequestInput,
    build_final_timesheet_report,
    export_to_final_timesheet,
)


def test_build_final_timesheet_report_applies_hr_rules_and_summary():
    report = build_final_timesheet_report(
        period_start=date(2026, 4, 23),
        period_end=date(2026, 5, 22),
        employees=[
            FinalTimesheetEmployeeInput(
                employee_id=1,
                machine_employee_id="E001",
                full_name="Nguyen Van A",
                department_name="Operations",
                previous_paid_leave_balance=2,
                current_month_paid_leave_credit=1,
            )
        ],
        daily_records=[
            FinalTimesheetDailyInput(
                employee_id=1,
                work_date=date(2026, 4, 23),
                attendance_symbol="X",
                check_in_time="08:30",
                check_out_time="17:30",
            ),
            FinalTimesheetDailyInput(
                employee_id=1,
                work_date=date(2026, 4, 25),
                attendance_symbol="V",
                check_in_time=None,
                check_out_time=None,
            ),
            FinalTimesheetDailyInput(
                employee_id=1,
                work_date=date(2026, 4, 27),
                attendance_symbol="X",
                check_in_time="08:30",
                check_out_time="17:30",
            ),
            FinalTimesheetDailyInput(
                employee_id=1,
                work_date=date(2026, 4, 28),
                attendance_symbol="V",
                check_in_time=None,
                check_out_time=None,
            ),
            FinalTimesheetDailyInput(
                employee_id=1,
                work_date=date(2026, 4, 29),
                attendance_symbol="X",
                check_in_time="08:15",
                check_out_time="17:15",
            ),
        ],
        entry_records=[
            FinalTimesheetEntryInput(
                employee_id=1,
                work_date=date(2026, 4, 29),
                final_symbol="P/X",
                check_in_time="08:15",
                check_out_time="17:15",
            )
        ],
        off_requests=[
            FinalTimesheetOffRequestInput(
                employee_id=1,
                request_type="paid_leave",
                start_date=date(2026, 4, 24),
                end_date=date(2026, 4, 24),
                total_days=1,
            ),
            FinalTimesheetOffRequestInput(
                employee_id=1,
                request_type="paid_leave_pm",
                start_date=date(2026, 4, 27),
                end_date=date(2026, 4, 27),
                total_days=0.5,
            ),
        ],
    )

    row = report.rows[0]
    assert row.days["2026-04-23"] == "X"
    assert row.days["2026-04-24"] == "P"
    assert row.days["2026-04-25"] == ""
    assert row.days["2026-04-26"] == ""
    assert row.days["2026-04-27"] == "X/P"
    assert row.days["2026-04-28"] == "Ro"
    assert row.days["2026-04-29"] == "P/X"
    assert row.total_work_days == 2.0
    assert row.paid_leave_days == 2.0
    assert row.unpaid_leave_days == 1.0
    # Ro là nghỉ không lương, không tiêu hao quỹ phép hưởng lương.
    assert row.remaining_paid_leave_days == 1.0


def test_accountant_payroll_days_are_preserved_separately_from_actual_days():
    report = build_final_timesheet_report(
        period_start=date(2026, 8, 20),
        period_end=date(2026, 8, 22),
        employees=[
            FinalTimesheetEmployeeInput(
                employee_id=1,
                machine_employee_id="8",
                full_name="Pham Do Hanh Quyen",
                stored_total_work_days=20,
                stored_total_payroll_days=23,
                stored_total_paid_leave_days=2.5,
                prefer_stored_totals=True,
            )
        ],
        daily_records=[],
    )

    row = report.rows[0]
    assert row.total_work_days == 20
    assert row.total_payroll_days == 23
    assert row.paid_leave_days == 2.5

    workbook = load_workbook(BytesIO(export_to_final_timesheet(report).getvalue()))
    sheet = workbook["Timesheet"]
    # A/B = identity, C:E = 3 dates, F = spacer, G/H = Ngày công/Ngày công TT.
    assert sheet["G8"].value == 23
    assert sheet["H8"].value == 20


def test_weekend_notion_leave_is_blank_and_continues_on_next_workday():
    report = build_final_timesheet_report(
        period_start=date(2026, 7, 3),
        period_end=date(2026, 7, 6),
        employees=[
            FinalTimesheetEmployeeInput(
                employee_id=1,
                machine_employee_id="E001",
                full_name="Nguyễn Thanh Đạt",
            )
        ],
        daily_records=[
            # Dữ liệu lịch trình cũ có X vào T7 nhưng không có quẹt thẻ.
            FinalTimesheetDailyInput(
                employee_id=1,
                work_date=date(2026, 7, 4),
                attendance_symbol="X",
            )
        ],
        off_requests=[
            FinalTimesheetOffRequestInput(
                employee_id=1,
                request_type="paid_leave",
                start_date=date(2026, 7, 3),
                end_date=date(2026, 7, 6),
                total_days=2,
            )
        ],
    )

    row = report.rows[0]
    assert row.days == {
        "2026-07-03": "P",
        "2026-07-04": "",
        "2026-07-05": "",
        "2026-07-06": "P",
    }
    assert row.paid_leave_days == 2.0
    assert row.total_work_days == 0.0

    workbook = load_workbook(BytesIO(export_to_final_timesheet(report).getvalue()))
    sheet = workbook["Timesheet"]
    assert sheet["B8"].value == "Nguyễn Thanh Đạt"
    assert sheet["C8"].value == 1.0
    assert sheet["D8"].value is None
    assert sheet["E8"].value is None
    assert sheet["F8"].value == 1.0


def test_pending_notion_leave_overrides_machine_absence_in_final_grid():
    report = build_final_timesheet_report(
        period_start=date(2026, 7, 6),
        period_end=date(2026, 7, 6),
        employees=[
            FinalTimesheetEmployeeInput(
                employee_id=26,
                machine_employee_id="26",
                full_name="Nguyen Thanh Dat",
            )
        ],
        daily_records=[
            FinalTimesheetDailyInput(
                employee_id=26,
                work_date=date(2026, 7, 6),
                attendance_symbol="V",
            )
        ],
        off_requests=[
            FinalTimesheetOffRequestInput(
                employee_id=26,
                request_type="paid_leave",
                start_date=date(2026, 7, 6),
                end_date=date(2026, 7, 6),
                total_days=1,
                status="pending",
            )
        ],
    )

    row = report.rows[0]
    assert row.days["2026-07-06"] == "P"
    assert row.total_absent_days == 0.0
    assert row.paid_leave_days == 1.0


def test_weekend_is_blank_even_with_punches_and_manual_override():
    report = build_final_timesheet_report(
        period_start=date(2026, 7, 4),
        period_end=date(2026, 7, 5),
        employees=[
            FinalTimesheetEmployeeInput(
                employee_id=1,
                machine_employee_id="29",
                full_name="Nguyễn Thị Thanh Hương",
            )
        ],
        daily_records=[
            FinalTimesheetDailyInput(
                employee_id=1,
                work_date=date(2026, 7, 4),
                attendance_symbol="X",
                check_in_time="08:10",
                check_out_time="17:40",
            )
        ],
        entry_records=[
            FinalTimesheetEntryInput(
                employee_id=1,
                work_date=date(2026, 7, 5),
                final_symbol="X",
                check_in_time="08:20",
                check_out_time="17:30",
                is_overridden=True,
            )
        ],
    )

    row = report.rows[0]
    assert row.days == {"2026-07-04": "", "2026-07-05": ""}
    assert row.total_work_days == 0


def test_work_symbol_without_machine_punch_is_payable_but_not_actual_attendance():
    report = build_final_timesheet_report(
        period_start=date(2026, 4, 23),
        period_end=date(2026, 5, 22),
        employees=[
            FinalTimesheetEmployeeInput(
                employee_id=1,
                machine_employee_id="E001",
                full_name="Nguyen Van A",
            )
        ],
        daily_records=[
            FinalTimesheetDailyInput(
                employee_id=1,
                work_date=date(2026, 4, 30),
                attendance_symbol="X",
                check_in_time=None,
                check_out_time=None,
            )
        ],
    )

    row = report.rows[0]
    assert row.days["2026-04-30"] == "X"
    assert row.total_work_days == 0.0
    assert row.total_payroll_days == 1.0


def test_payroll_days_include_clocked_work_paid_leave_and_wfh():
    report = build_final_timesheet_report(
        period_start=date(2026, 8, 3),
        period_end=date(2026, 8, 5),
        employees=[
            FinalTimesheetEmployeeInput(
                employee_id=1,
                machine_employee_id="E001",
                full_name="Nguyen Van A",
            )
        ],
        daily_records=[
            FinalTimesheetDailyInput(
                employee_id=1,
                work_date=date(2026, 8, 3),
                attendance_symbol="X",
                check_in_time="08:30",
                check_out_time="17:30",
            ),
            # WFH is stored as X without machine punches.
            FinalTimesheetDailyInput(
                employee_id=1,
                work_date=date(2026, 8, 4),
                attendance_symbol="X",
            ),
        ],
        off_requests=[
            FinalTimesheetOffRequestInput(
                employee_id=1,
                request_type="paid_leave",
                start_date=date(2026, 8, 5),
                end_date=date(2026, 8, 5),
                total_days=1,
            )
        ],
    )

    row = report.rows[0]
    assert row.total_work_days == 1.0
    assert row.paid_leave_days == 1.0
    assert row.total_payroll_days == 3.0


def test_build_final_timesheet_report_marks_single_punch_day_as_worked():
    report = build_final_timesheet_report(
        period_start=date(2026, 4, 23),
        period_end=date(2026, 5, 22),
        employees=[
            FinalTimesheetEmployeeInput(
                employee_id=1,
                machine_employee_id="E001",
                full_name="Nguyen Van A",
            )
        ],
        daily_records=[
            FinalTimesheetDailyInput(
                employee_id=1,
                work_date=date(2026, 4, 23),
                attendance_symbol=None,
                check_in_time="17:43",
                check_out_time=None,
                abnormal_level="L1",
            )
        ],
    )

    row = report.rows[0]
    assert row.days["2026-04-23"] == "X"
    assert row.total_work_days == 1.0


def test_export_to_final_timesheet_creates_hr_template_with_border():
    report = build_final_timesheet_report(
        period_start=date(2026, 4, 23),
        period_end=date(2026, 5, 22),
        employees=[
            FinalTimesheetEmployeeInput(
                employee_id=1,
                machine_employee_id="E001",
                full_name="Nguyen Van A",
                previous_paid_leave_balance=2,
                current_month_paid_leave_credit=1,
            )
        ],
        daily_records=[
            FinalTimesheetDailyInput(
                employee_id=1,
                work_date=date(2026, 4, 23),
                attendance_symbol="X",
                check_in_time="08:30",
                check_out_time="17:30",
            )
        ],
        off_requests=[
            FinalTimesheetOffRequestInput(
                employee_id=1,
                request_type="paid_leave",
                start_date=date(2026, 4, 24),
                end_date=date(2026, 4, 24),
                total_days=1,
            )
        ],
    )

    output = export_to_final_timesheet(report)
    workbook = load_workbook(BytesIO(output.getvalue()))
    sheet = workbook["Timesheet"]
    assert sheet["G6"].value == "EMPLOYEE TIMESHEET FROM 23/04/2026 TO 22/05/2026"
    assert sheet["C7"].value == 23
    assert sheet["E7"].value == "25\nT7"
    assert sheet["C8"].value == 1.0
    assert sheet["D8"].value == 1.0
    assert sheet["AK8"].value == "X"
    assert sheet["AL8"].value == "P"
    assert sheet["BO8"].value == 0
    assert sheet["BP8"].value == 1
    assert sheet["BQ8"].value == 2
    assert sheet["BR8"].value == 1
    assert sheet["BS8"].value == 2
    assert sheet["C8"].number_format == "0.0"
    assert sheet["BO8"].number_format == "0.0"
    assert sheet["C8"].border.left.style == "thin"


def test_export_to_final_timesheet_formats_half_day_work_units_with_one_decimal():
    report = build_final_timesheet_report(
        period_start=date(2026, 4, 23),
        period_end=date(2026, 5, 22),
        employees=[
            FinalTimesheetEmployeeInput(
                employee_id=1,
                machine_employee_id="E001",
                full_name="Nguyen Van A",
            )
        ],
        daily_records=[],
        entry_records=[
            FinalTimesheetEntryInput(
                employee_id=1,
                work_date=date(2026, 4, 23),
                final_symbol="P/V",
            )
        ],
    )

    workbook = load_workbook(BytesIO(export_to_final_timesheet(report).getvalue()))
    cell = workbook["Timesheet"]["C8"]
    assert cell.value == 0.5
    assert cell.number_format == "0.0"


def test_build_final_timesheet_report_half_day_leave_with_units_and_attendance():
    # Similar to BOO BAO's case: 1.0 day units in request, afternoon time range (e.g. 13:30 -> 17:30), check-out at 13:16
    report = build_final_timesheet_report(
        period_start=date(2026, 6, 1),
        period_end=date(2026, 6, 30),
        employees=[
            FinalTimesheetEmployeeInput(
                employee_id=4,
                machine_employee_id="20",
                full_name="Gia Bao",
            )
        ],
        daily_records=[
            # Check-out at 13:16 is before 13:30, so no afternoon work
            FinalTimesheetDailyInput(
                employee_id=4,
                work_date=date(2026, 6, 19),
                attendance_symbol="X",
                check_in_time="08:51",
                check_out_time="13:16",
            )
        ],
        off_requests=[
            # request_type contains pm indicating afternoon, total_days might be 1.0 or 0.5
            FinalTimesheetOffRequestInput(
                employee_id=4,
                request_type="paid_leave_pm",
                start_date=date(2026, 6, 19),
                end_date=date(2026, 6, 19),
                total_days=1.0,
            )
        ],
    )

    row = report.rows[0]
    # Afternoon leave + morning work + check-out at 13:16 (no afternoon work) should resolve to X/P
    assert row.days["2026-06-19"] == "X/P"
    assert row.total_work_days == 0.5
    assert row.paid_leave_days == 0.5
    assert row.unpaid_leave_days == 0.0


def test_paid_leave_balance_rolls_forward_from_july_to_august_without_deducting_unpaid_leave():
    july_report = build_final_timesheet_report(
        period_start=date(2026, 6, 23),
        period_end=date(2026, 7, 22),
        daily_records=[],
        employees=[
            FinalTimesheetEmployeeInput(
                employee_id=1,
                machine_employee_id="E001",
                full_name="Nguyen Van A",
                previous_paid_leave_balance=5,
                current_month_paid_leave_credit=1,
            )
        ],
        off_requests=[
            FinalTimesheetOffRequestInput(
                employee_id=1,
                request_type="paid_leave",
                start_date=date(2026, 7, 6),
                end_date=date(2026, 7, 6),
                total_days=1,
            ),
            FinalTimesheetOffRequestInput(
                employee_id=1,
                request_type="unpaid_leave",
                start_date=date(2026, 7, 7),
                end_date=date(2026, 7, 7),
                total_days=1,
            ),
        ],
    )
    july_row = july_report.rows[0]
    assert july_row.paid_leave_days == 1
    assert july_row.unpaid_leave_days == 1
    assert july_row.remaining_paid_leave_days == 5

    august_report = build_final_timesheet_report(
        period_start=date(2026, 7, 23),
        period_end=date(2026, 8, 22),
        daily_records=[],
        employees=[
            FinalTimesheetEmployeeInput(
                employee_id=1,
                machine_employee_id="E001",
                full_name="Nguyen Van A",
                previous_paid_leave_balance=july_row.remaining_paid_leave_days,
                current_month_paid_leave_credit=1,
            )
        ],
        off_requests=[
            FinalTimesheetOffRequestInput(
                employee_id=1,
                request_type="paid_leave",
                start_date=date(2026, 8, 3),
                end_date=date(2026, 8, 3),
                total_days=0.5,
            )
        ],
    )
    august_row = august_report.rows[0]
    assert august_row.previous_paid_leave_balance == 5
    assert august_row.paid_leave_days == 0.5
    assert august_row.remaining_paid_leave_days == 5.5


def test_single_saturday_working_day_override_does_not_change_other_weekends():
    report = build_final_timesheet_report(
        period_start=date(2026, 8, 15),
        period_end=date(2026, 8, 22),
        employees=[
            FinalTimesheetEmployeeInput(
                employee_id=1,
                machine_employee_id="E001",
                full_name="Nguyen Van A",
            )
        ],
        daily_records=[
            FinalTimesheetDailyInput(
                employee_id=1,
                work_date=date(2026, 8, 15),
                attendance_symbol="X",
            ),
            FinalTimesheetDailyInput(
                employee_id=1,
                work_date=date(2026, 8, 22),
                attendance_symbol="X",
                check_in_time="08:30",
                check_out_time="17:30",
            ),
        ],
        working_day_overrides={date(2026, 8, 22)},
    )

    row = report.rows[0]
    assert row.days["2026-08-15"] == ""
    assert row.days["2026-08-22"] == "X"
    assert row.total_work_days == 1


def test_accountant_blank_override_stays_blank_on_a_weekday():
    report = build_final_timesheet_report(
        period_start=date(2026, 8, 3),
        period_end=date(2026, 8, 3),
        employees=[
            FinalTimesheetEmployeeInput(
                employee_id=1,
                machine_employee_id="E001",
                full_name="New Employee",
            )
        ],
        daily_records=[
            FinalTimesheetDailyInput(
                employee_id=1,
                work_date=date(2026, 8, 3),
                attendance_symbol="",
            )
        ],
        entry_records=[
            FinalTimesheetEntryInput(
                employee_id=1,
                work_date=date(2026, 8, 3),
                final_symbol="",
                is_overridden=True,
                override_reason="Kế toán duyệt để trống",
            )
        ],
    )

    assert report.rows[0].days["2026-08-03"] == ""

