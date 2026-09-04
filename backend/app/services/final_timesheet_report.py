from __future__ import annotations

from dataclasses import dataclass
from datetime import date, timedelta
from io import BytesIO
from typing import Any, Iterable

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import and_
from sqlalchemy.orm import Session

from app.models.attendance_daily import AttendanceDaily
from app.models.employee import Employee
from app.models.holiday_setting import HolidaySetting
from app.models.off_request import OffRequest
from app.models.timesheet import Timesheet
from app.models.timesheet_entry import TimesheetEntry
from app.models.timesheet_period import TimesheetPeriod


WEEKDAY_LABELS = {
    0: "T2",
    1: "T3",
    2: "T4",
    3: "T5",
    4: "T6",
    5: "T7",
    6: "CN",
}

# A submitted Notion leave is provisional, but it is still a valid attendance
# exception.  This prevents a submitted request from being rendered as V while
# approval is pending.  Rejected/cancelled requests are deliberately excluded.
ACTIVE_OFF_REQUEST_STATUSES = {"approved", "approve", "pending", "submitted", "under review"}
ACTIVE_OFF_REQUEST_DB_STATUSES = sorted(ACTIVE_OFF_REQUEST_STATUSES | {"under_review"})

THIN_GREY_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


@dataclass
class FinalTimesheetEmployeeInput:
    employee_id: int
    machine_employee_id: str
    full_name: str
    department_name: str | None = None
    employee_type: str = "FULLTIME"
    previous_paid_leave_balance: float = 0.0
    current_month_paid_leave_credit: float = 0.0
    stored_total_work_days: float | None = None
    stored_total_payroll_days: float | None = None
    stored_total_paid_leave_days: float | None = None
    stored_total_unpaid_leave_days: float | None = None
    stored_total_absent_days: float | None = None
    stored_total_late_minutes: int | None = None
    stored_remaining_paid_leave_days: float | None = None
    prefer_stored_totals: bool = False
    preserve_leave_snapshot: bool = False


@dataclass
class FinalTimesheetDailyInput:
    employee_id: int
    work_date: date
    attendance_symbol: str | None = None
    check_in_time: str | None = None
    check_out_time: str | None = None
    late_minutes: int = 0
    early_minutes: int = 0
    abnormal_level: str | None = None


@dataclass
class FinalTimesheetEntryInput:
    employee_id: int
    work_date: date
    final_symbol: str
    check_in_time: str | None = None
    check_out_time: str | None = None
    is_overridden: bool = False
    override_reason: str | None = None


@dataclass
class FinalTimesheetOffRequestInput:
    employee_id: int
    request_type: str
    start_date: date
    end_date: date
    total_days: float
    status: str = "approved"


@dataclass
class FinalTimesheetDayColumn:
    key: str
    day_number: int
    weekday_label: str
    is_weekend: bool


@dataclass
class FinalTimesheetRow:
    employee_id: int
    machine_employee_id: str
    full_name: str
    department_name: str | None
    days: dict[str, str]
    override_reasons: dict[str, str]
    abnormal_days: int
    total_late_minutes: int
    total_early_minutes: int
    total_absent_days: float
    total_work_days: float
    total_payroll_days: float
    unpaid_leave_days: float
    paid_leave_days: float
    previous_paid_leave_balance: float
    current_month_paid_leave_credit: float
    remaining_paid_leave_days: float


@dataclass
class FinalTimesheetReport:
    period_start: date
    period_end: date
    day_keys: list[str]
    day_columns: list[FinalTimesheetDayColumn]
    rows: list[FinalTimesheetRow]


def _round_leave(value: float) -> float:
    return round(float(value or 0), 2)


def _normalize_text(value: str | None) -> str:
    return " ".join(str(value or "").strip().lower().replace("_", " ").replace("-", " ").split())


def _clean_symbol(value: str | None) -> str:
    text = str(value or "").strip().upper().replace(" ", "")
    if text in {"XP", "X-P", "X\\P"}:
        return "X/P"
    if text in {"PX", "P-X", "P\\X"}:
        return "P/X"
    # `Ro` is the accounting template's canonical code for unpaid absence.
    # Keep V/O aliases so historic imports continue to render correctly.
    if text in {"PV", "P-V", "P\\V", "P/V", "P/RO", "PRO", "P-RO", "P\\RO"}:
        return "P/Ro"
    if text in {"VP", "V-P", "V\\P", "V/P", "RO/P", "ROP", "RO-P", "RO\\P"}:
        return "Ro/P"
    if text in {"XV", "X-V", "X\\V", "X/V", "X/RO", "XRO", "X-RO", "X\\RO"}:
        return "X/Ro"
    if text in {"VX", "V-X", "V\\X", "V/X", "RO/X", "ROX", "RO-X", "RO\\X"}:
        return "Ro/X"
    if text in {"O", "V", "RO"}:
        return "Ro"
    if text in {"X/P", "P/X", "P/RO", "RO/P", "X/RO", "RO/X", "X", "P", "CT"}:
        return {"P/RO": "P/Ro", "RO/P": "Ro/P", "X/RO": "X/Ro", "RO/X": "Ro/X"}.get(text, text)
    return text


def _iter_period_dates(period_start: date, period_end: date) -> list[date]:
    days: list[date] = []
    cursor = period_start
    while cursor <= period_end:
        days.append(cursor)
        cursor += timedelta(days=1)
    return days


def _build_day_columns(
    period_start: date,
    period_end: date,
    working_day_overrides: set[date] | None = None,
) -> list[FinalTimesheetDayColumn]:
    working_day_overrides = working_day_overrides or set()
    return [
        FinalTimesheetDayColumn(
            key=work_date.isoformat(),
            day_number=work_date.day,
            weekday_label=WEEKDAY_LABELS[work_date.weekday()],
            is_weekend=work_date.weekday() >= 5 and work_date not in working_day_overrides,
        )
        for work_date in _iter_period_dates(period_start, period_end)
    ]


def _resolve_half_day_symbol(request_type: str) -> str:
    normalized = _normalize_text(request_type)
    if "p/x" in normalized or "morning" in normalized or "am" in normalized or "sang" in normalized:
        return "P/X"
    if "x/p" in normalized or "afternoon" in normalized or "pm" in normalized or "chieu" in normalized:
        return "X/P"
    return "X/P"


def _resolve_half_day_unpaid_symbol(request_type: str) -> str:
    normalized = _normalize_text(request_type)
    if "v/x" in normalized or "morning" in normalized or "am" in normalized or "sang" in normalized:
        return "Ro/X"
    if "x/v" in normalized or "afternoon" in normalized or "pm" in normalized or "chieu" in normalized:
        return "X/Ro"
    return "X/Ro"


def _is_business_trip(request_type: str) -> bool:
    normalized = _normalize_text(request_type)
    return normalized in {"ct", "business trip"} or "business trip" in normalized or "cong tac" in normalized


def _is_unpaid_leave(request_type: str) -> bool:
    normalized = _normalize_text(request_type)
    return "unpaid" in normalized or "khong luong" in normalized


def _is_paid_leave(request_type: str) -> bool:
    normalized = _normalize_text(request_type)
    if _is_unpaid_leave(request_type):
        return False
    return (
        normalized in {"p", "paid", "paid leave", "annual leave"}
        or "paid" in normalized
        or "phep" in normalized
    )


def _resolve_leave_day_detail(request_type: str, units: float) -> dict[str, Any]:
    normalized = _normalize_text(request_type)
    
    if _is_business_trip(request_type):
        return {
            "leave_type": "business",
            "session": "full",
            "display_symbol": "CT",
            "valid_leave": True
        }
        
    is_half_day = any(marker in normalized for marker in ["am", "pm", "morning", "afternoon", "sang", "chieu"])
    session = "full"
    if is_half_day:
        if any(marker in normalized for marker in ["morning", "am", "sang", "p/x", "v/x"]):
            session = "morning"
        elif any(marker in normalized for marker in ["afternoon", "pm", "chieu", "x/p", "x/v"]):
            session = "afternoon"
    else:
        if units < 1:
            session = "afternoon"

    if _is_unpaid_leave(request_type):
        if session == "morning":
            display_symbol = "Ro/X"
        elif session == "afternoon":
            display_symbol = "X/Ro"
        else:
            display_symbol = "Ro"
        return {
            "leave_type": "unpaid",
            "session": session,
            "display_symbol": display_symbol,
            "valid_leave": True
        }
        
    if _is_paid_leave(request_type):
        if session == "morning":
            display_symbol = "P/X"
        elif session == "afternoon":
            display_symbol = "X/P"
        else:
            if units >= 1 and not is_half_day:
                display_symbol = "P"
            else:
                display_symbol = _resolve_half_day_symbol(request_type)
                # determine session from half day symbol
                if display_symbol == "P/X":
                    session = "morning"
                else:
                    session = "afternoon"
        return {
            "leave_type": "paid",
            "session": session,
            "display_symbol": display_symbol,
            "valid_leave": True
        }
        
    return {
        "leave_type": "unknown",
        "session": "full",
        "display_symbol": "",
        "valid_leave": True
    }


def _build_off_request_map(
    off_requests: Iterable[FinalTimesheetOffRequestInput],
    period_start: date,
    period_end: date,
) -> dict[tuple[int, date], dict[str, Any]]:
    off_request_map: dict[tuple[int, date], dict[str, Any]] = {}
    for request in off_requests:
        if _normalize_text(request.status) not in ACTIVE_OFF_REQUEST_STATUSES:
            continue
        remaining_days = float(request.total_days or 0)
        if remaining_days <= 0:
            continue
        cursor = max(request.start_date, period_start)
        request_end = min(request.end_date, period_end)
        while cursor <= request_end and remaining_days > 0:
            # T7/CN là ngày nghỉ mặc định. Đơn Notion có khoảng ngày bao gồm
            # cuối tuần không được tạo ký hiệu phép và cũng không tiêu hao số
            # ngày của đơn; phần còn lại tiếp tục được phân bổ vào ngày làm việc.
            if cursor.weekday() >= 5:
                cursor += timedelta(days=1)
                continue
            units = 1.0 if remaining_days >= 1 else 0.5
            off_request_map[(request.employee_id, cursor)] = _resolve_leave_day_detail(request.request_type, units)
            remaining_days = round(remaining_days - units, 2)
            cursor += timedelta(days=1)
    return off_request_map


def _work_units_for_symbol(symbol: str) -> float:
    """Return payable work units represented by an attendance symbol.

    This intentionally includes WFH/business-trip days represented as
    ``X``/``CT`` even when there is no fingerprint-machine punch. It is used
    to derive payroll days, not actual clocked days.
    """
    symbol = _clean_symbol(symbol)
    if symbol in {"X", "CT"}:
        return 1.0
    if symbol in {"X/P", "P/X", "X/Ro", "Ro/X"}:
        return 0.5
    return 0.0


def _clocked_work_units_for_symbol(
    symbol: str,
    check_in_time: str | None,
    check_out_time: str | None,
) -> float:
    """Return actual work units backed by at least one machine punch."""
    if not _effective_time(check_in_time, check_out_time):
        return 0.0
    return _work_units_for_symbol(symbol)


def _paid_leave_units_for_symbol(symbol: str) -> float:
    symbol = _clean_symbol(symbol)
    if symbol == "P":
        return 1.0
    if symbol in {"X/P", "P/X", "P/Ro", "Ro/P"}:
        return 0.5
    return 0.0


def _absent_units_for_symbol(symbol: str) -> float:
    symbol = _clean_symbol(symbol)
    if symbol == "Ro":
        return 1.0
    if symbol in {"P/Ro", "Ro/P", "X/Ro", "Ro/X"}:
        return 0.5
    return 0.0


def _effective_time(entry_value: str | None, daily_value: str | None) -> str | None:
    for candidate in [entry_value, daily_value]:
        if candidate is not None and str(candidate).strip():
            return str(candidate).strip()
    return None


def _resolve_display_symbol(
    column: FinalTimesheetDayColumn,
    daily: FinalTimesheetDailyInput | None,
    entry: FinalTimesheetEntryInput | None,
    leave_detail: dict[str, Any] | None,
) -> tuple[str, bool, bool]:
    # T7/CN là ngày nghỉ mặc định tuyệt đối trong bảng công. Dữ liệu quẹt thẻ
    # vẫn được giữ ở bảng log để đối soát, nhưng không được tạo ký hiệu/công,
    # kể cả khi có quẹt thẻ hoặc từng có override thủ công.
    if column.is_weekend:
        return "", False, False

    # 1. Admin manual override takes ultimate precedence
    if entry and getattr(entry, "is_overridden", False):
        explicit_symbol = _clean_symbol(entry.final_symbol)
        # An accountant-approved blank is also an explicit decision (for
        # example dates before a new employee starts). Do not turn it back
        # into Ro merely because an AttendanceDaily placeholder exists.
        return explicit_symbol, bool(leave_detail and leave_detail.get("valid_leave")), bool(explicit_symbol)

    check_in_time = _effective_time(entry.check_in_time if entry else None, daily.check_in_time if daily else None)
    check_out_time = _effective_time(entry.check_out_time if entry else None, daily.check_out_time if daily else None)
    has_any_attendance = bool(check_in_time or check_out_time)
    has_full_attendance = bool(check_in_time and check_out_time)
    valid_leave = bool(leave_detail and leave_detail.get("valid_leave"))
    raw_symbol = _clean_symbol(daily.attendance_symbol if daily else None)

    # 2. If daily record already has a reconciled compound symbol, use it
    if raw_symbol in {"X/P", "P/X", "P/Ro", "Ro/P", "X/Ro", "Ro/X"}:
        return raw_symbol, valid_leave, has_any_attendance

    # 3. If there is a valid leave request from the DB and it is not manually overridden, resolve the symbol
    if leave_detail and leave_detail.get("valid_leave"):
        ltype = leave_detail.get("leave_type")
        session = leave_detail.get("session", "full")
        
        # Parse check-in / check-out times to minutes
        in_mins = None
        if check_in_time:
            try:
                h, m = map(int, check_in_time.split(":"))
                in_mins = h * 60 + m
            except Exception:
                pass
        out_mins = None
        if check_out_time:
            try:
                h, m = map(int, check_out_time.split(":"))
                out_mins = h * 60 + m
            except Exception:
                pass

        # Unified threshold: afternoon work starts at 13:30 (so checkout must be >= 13:30)
        has_morning_work = in_mins is not None and in_mins <= 12 * 60
        has_afternoon_work = out_mins is not None and out_mins >= 13 * 60 + 30

        if ltype == "business":
            return "CT", valid_leave, has_any_attendance
            
        elif ltype in {"paid", "unpaid"}:
            leave_sym = "Ro" if ltype == "unpaid" else "P"
            
            if session == "morning":
                am_sym = leave_sym
                pm_sym = "X" if has_afternoon_work else "Ro"
            elif session == "afternoon":
                am_sym = "X" if has_morning_work else "Ro"
                pm_sym = leave_sym
            else:
                am_sym = "X" if has_morning_work else leave_sym
                pm_sym = "X" if has_afternoon_work else leave_sym
                
            display_symbol = am_sym if am_sym == pm_sym else f"{am_sym}/{pm_sym}"
            return display_symbol, valid_leave, has_any_attendance

    # 4. Fallback to entry.final_symbol if it exists
    explicit_symbol = _clean_symbol(entry.final_symbol if entry else None)
    if explicit_symbol:
        return explicit_symbol, valid_leave, has_any_attendance

    # 5. General logic fallback (for when there's no entry)
    if has_full_attendance:
        return "X", valid_leave, True

    if has_any_attendance:
        return "X", valid_leave, True

    if column.is_weekend and not has_any_attendance and raw_symbol in {"", "Ro"} and not valid_leave:
        return "", False, False

    if raw_symbol:
        return raw_symbol, valid_leave, has_any_attendance

    if not has_any_attendance and not valid_leave and column.is_weekend and daily is None and entry is None:
        return "", False, False

    if daily is not None and not has_any_attendance and not valid_leave:
        return "Ro", False, False

    return "", valid_leave, has_any_attendance


def build_final_timesheet_report(
    period_start: date,
    period_end: date,
    employees: Iterable[FinalTimesheetEmployeeInput],
    daily_records: Iterable[FinalTimesheetDailyInput],
    entry_records: Iterable[FinalTimesheetEntryInput] | None = None,
    off_requests: Iterable[FinalTimesheetOffRequestInput] | None = None,
    working_day_overrides: set[date] | None = None,
) -> FinalTimesheetReport:
    day_columns = _build_day_columns(period_start, period_end, working_day_overrides)
    day_keys = [column.key for column in day_columns]
    daily_map = {(row.employee_id, row.work_date): row for row in daily_records}
    entry_map = {(row.employee_id, row.work_date): row for row in (entry_records or [])}
    off_request_map = _build_off_request_map(off_requests or [], period_start, period_end)

    rows: list[FinalTimesheetRow] = []
    for employee in employees:
        day_symbols = {key: "" for key in day_keys}
        override_reasons = {}
        abnormal_days = 0
        total_late_minutes = 0
        total_early_minutes = 0
        total_absent_days = 0.0
        total_work_days = 0.0
        payable_work_days = 0.0
        unpaid_leave_days = 0.0
        paid_leave_days = 0.0
        has_detail = False

        for column in day_columns:
            work_date = date.fromisoformat(column.key)
            daily = daily_map.get((employee.employee_id, work_date))
            entry = entry_map.get((employee.employee_id, work_date))
            leave_detail = off_request_map.get((employee.employee_id, work_date))
            # Học việc/thử việc không có phép năm. Một đơn được gửi dưới dạng
            # "nghỉ phép" vẫn phải hiện Ro và được tính là nghỉ không lương.
            if (
                leave_detail
                and str(employee.employee_type or "FULLTIME").upper() != "FULLTIME"
                and leave_detail.get("leave_type") == "paid"
            ):
                leave_detail = {**leave_detail, "leave_type": "unpaid"}
            display_symbol, valid_leave, has_any_attendance = _resolve_display_symbol(column, daily, entry, leave_detail)
            day_symbols[column.key] = display_symbol
            if entry and entry.is_overridden and entry.override_reason:
                override_reasons[column.key] = entry.override_reason

            late_minutes = int(daily.late_minutes if daily else 0)
            early_minutes = int(daily.early_minutes if daily else 0)
            check_in_time = _effective_time(
                entry.check_in_time if entry else None,
                daily.check_in_time if daily else None,
            )
            check_out_time = _effective_time(
                entry.check_out_time if entry else None,
                daily.check_out_time if daily else None,
            )
            total_late_minutes += late_minutes
            total_early_minutes += early_minutes
            total_work_days += _clocked_work_units_for_symbol(display_symbol, check_in_time, check_out_time)
            payable_work_days += _work_units_for_symbol(display_symbol)
            paid_leave_days += _paid_leave_units_for_symbol(display_symbol)
            absent_units = _absent_units_for_symbol(display_symbol)
            unpaid_leave_days += absent_units
            total_absent_days += absent_units

            is_unpaid_leave_day = bool(
                daily is not None
                and not valid_leave
                and not has_any_attendance
                and display_symbol == "Ro"
            )
            if is_unpaid_leave_day:
                unpaid_leave_days += 1.0 - absent_units
                total_absent_days += 1.0 - absent_units

            if bool(daily and daily.abnormal_level) or late_minutes > 0 or early_minutes > 0 or (has_any_attendance and display_symbol == ""):
                abnormal_days += 1

            if daily is not None or entry is not None or leave_detail is not None or display_symbol:
                has_detail = True

        if not has_detail or employee.prefer_stored_totals:
            if employee.stored_total_work_days is not None:
                total_work_days = _round_leave(employee.stored_total_work_days)
                if not has_detail:
                    payable_work_days = total_work_days
            if employee.stored_total_paid_leave_days is not None:
                paid_leave_days = _round_leave(employee.stored_total_paid_leave_days)
            if employee.stored_total_unpaid_leave_days is not None:
                unpaid_leave_days = _round_leave(employee.stored_total_unpaid_leave_days)
            if employee.stored_total_absent_days is not None:
                total_absent_days = _round_leave(employee.stored_total_absent_days)
            if employee.stored_total_late_minutes is not None:
                total_late_minutes = int(employee.stored_total_late_minutes)

        total_payroll_days = _round_leave(
            employee.stored_total_payroll_days
            if employee.stored_total_payroll_days is not None
            else payable_work_days + paid_leave_days
        )

        is_fulltime = str(employee.employee_type or "FULLTIME").upper() == "FULLTIME"
        previous_paid_leave_balance = _round_leave(employee.previous_paid_leave_balance)
        current_month_paid_leave_credit = _round_leave(
            employee.current_month_paid_leave_credit
            if is_fulltime or employee.preserve_leave_snapshot
            else 0
        )
        # Chỉ phép hưởng lương tiêu hao quỹ phép năm. Nghỉ không lương (Ro)
        # được theo dõi riêng và không được làm giảm số dư phép hưởng lương.
        remaining_paid_leave_days = (
            _round_leave(employee.stored_remaining_paid_leave_days)
            if employee.preserve_leave_snapshot and employee.stored_remaining_paid_leave_days is not None
            else _round_leave(
                max(0.0, previous_paid_leave_balance + current_month_paid_leave_credit - paid_leave_days)
            )
        )

        rows.append(
            FinalTimesheetRow(
                employee_id=employee.employee_id,
                machine_employee_id=employee.machine_employee_id,
                full_name=employee.full_name,
                department_name=employee.department_name,
                days=day_symbols,
                override_reasons=override_reasons,
                abnormal_days=abnormal_days,
                total_late_minutes=total_late_minutes,
                total_early_minutes=total_early_minutes,
                total_absent_days=_round_leave(total_absent_days),
                total_work_days=_round_leave(total_work_days),
                total_payroll_days=total_payroll_days,
                unpaid_leave_days=_round_leave(unpaid_leave_days),
                paid_leave_days=_round_leave(paid_leave_days),
                previous_paid_leave_balance=previous_paid_leave_balance,
                current_month_paid_leave_credit=current_month_paid_leave_credit,
                remaining_paid_leave_days=remaining_paid_leave_days,
            )
        )

    return FinalTimesheetReport(
        period_start=period_start,
        period_end=period_end,
        day_keys=day_keys,
        day_columns=day_columns,
        rows=rows,
    )


def build_final_timesheet_report_from_db(db: Session, period_start: date, period_end: date) -> FinalTimesheetReport:
    period_timesheets = (
        db.query(Timesheet)
        .filter(Timesheet.period_start == period_start, Timesheet.period_end == period_end)
        .order_by(Timesheet.employee_id.asc())
        .all()
    )
    excluded_employee_ids = {
        row.employee_id for row in period_timesheets if str(row.approval_status or "").lower() == "excluded"
    }
    timesheets = [row for row in period_timesheets if row.employee_id not in excluded_employee_ids]
    daily_rows = (
        db.query(AttendanceDaily)
        .filter(AttendanceDaily.period_start == period_start, AttendanceDaily.period_end == period_end)
        .order_by(AttendanceDaily.employee_id.asc(), AttendanceDaily.work_date.asc())
        .all()
    )
    entry_rows = (
        db.query(TimesheetEntry)
        .filter(TimesheetEntry.work_date >= period_start, TimesheetEntry.work_date <= period_end)
        .order_by(TimesheetEntry.employee_id.asc(), TimesheetEntry.work_date.asc(), TimesheetEntry.id.asc())
        .all()
    )
    off_requests = (
        db.query(OffRequest)
        .filter(
            OffRequest.status.in_(ACTIVE_OFF_REQUEST_DB_STATUSES),
            OffRequest.start_date <= period_end,
            OffRequest.end_date >= period_start,
        )
        .order_by(OffRequest.employee_id.asc(), OffRequest.start_date.asc(), OffRequest.id.asc())
        .all()
    )

    employee_ids = {row.employee_id for row in daily_rows}
    employee_ids.update(row.employee_id for row in entry_rows)
    employee_ids.update(row.employee_id for row in off_requests)
    employee_ids.update(row.employee_id for row in timesheets)
    employee_ids.difference_update(excluded_employee_ids)

    employees: list[Employee] = []
    if employee_ids:
        employees = (
            db.query(Employee)
            .filter(Employee.id.in_(sorted(employee_ids)))
            .order_by(Employee.id.asc())
            .all()
        )

    timesheet_map = {row.employee_id: row for row in timesheets}
    def monthly_paid_leave_credit(employee: Employee) -> float:
        if str(employee.employee_type or "FULLTIME").upper() != "FULLTIME":
            return 0.0
        return round(float(employee.annual_leave_quota or 0) / 12, 2)

    previous_balance_by_employee = {
        employee.id: float(employee.paid_leave_balance or 0)
        for employee in employees
    }
    if employee_ids:
        prior_locked_timesheets = (
            db.query(Timesheet)
            .join(
                TimesheetPeriod,
                and_(
                    TimesheetPeriod.period_start == Timesheet.period_start,
                    TimesheetPeriod.period_end == Timesheet.period_end,
                ),
            )
            .filter(
                TimesheetPeriod.is_locked.is_(True),
                Timesheet.period_end < period_start,
                Timesheet.employee_id.in_(sorted(employee_ids)),
            )
            .order_by(Timesheet.employee_id.asc(), Timesheet.period_end.asc(), Timesheet.id.asc())
            .all()
        )
        employees_by_id = {employee.id: employee for employee in employees}
        for prior in prior_locked_timesheets:
            employee = employees_by_id.get(prior.employee_id)
            if employee is None:
                continue
            previous_balance = previous_balance_by_employee[prior.employee_id]
            if prior.remaining_paid_leave_days is not None:
                previous_balance_by_employee[prior.employee_id] = float(prior.remaining_paid_leave_days)
                continue
            prior_opening = (
                float(prior.previous_paid_leave_balance)
                if prior.previous_paid_leave_balance is not None
                else previous_balance
            )
            prior_credit = (
                float(prior.current_month_paid_leave_credit)
                if prior.current_month_paid_leave_credit is not None
                else monthly_paid_leave_credit(employee)
            )
            previous_balance_by_employee[prior.employee_id] = max(
                0.0,
                prior_opening + prior_credit - float(prior.total_paid_leave_days or 0),
            )

    def previous_balance_for(employee: Employee) -> float:
        current = timesheet_map.get(employee.id)
        if current and current.previous_paid_leave_balance is not None:
            return float(current.previous_paid_leave_balance)
        return previous_balance_by_employee.get(employee.id, 0.0)

    def current_credit_for(employee: Employee) -> float:
        current = timesheet_map.get(employee.id)
        if current and current.current_month_paid_leave_credit is not None:
            return float(current.current_month_paid_leave_credit)
        return monthly_paid_leave_credit(employee)

    employee_inputs = [
        FinalTimesheetEmployeeInput(
            employee_id=employee.id,
            machine_employee_id=employee.machine_employee_id,
            # Notion name is only a reconciliation key; all reports must show
            # the Vietnamese name maintained in the employee profile.
            full_name=employee.full_name,
            department_name=employee.department_name,
            employee_type=employee.employee_type,
            previous_paid_leave_balance=previous_balance_for(employee),
            current_month_paid_leave_credit=current_credit_for(employee),
            stored_total_work_days=float(timesheet_map[employee.id].total_work_days) if employee.id in timesheet_map else None,
            stored_total_payroll_days=(
                float(timesheet_map[employee.id].total_payroll_days)
                if employee.id in timesheet_map and timesheet_map[employee.id].total_payroll_days is not None
                else None
            ),
            stored_total_paid_leave_days=float(timesheet_map[employee.id].total_paid_leave_days) if employee.id in timesheet_map else None,
            stored_total_unpaid_leave_days=float(timesheet_map[employee.id].total_unpaid_leave_days) if employee.id in timesheet_map else None,
            stored_total_absent_days=float(timesheet_map[employee.id].total_absent_days) if employee.id in timesheet_map else None,
            stored_total_late_minutes=int(timesheet_map[employee.id].total_late_minutes) if employee.id in timesheet_map else None,
            stored_remaining_paid_leave_days=(
                float(timesheet_map[employee.id].remaining_paid_leave_days)
                if employee.id in timesheet_map and timesheet_map[employee.id].remaining_paid_leave_days is not None
                else None
            ),
            prefer_stored_totals=(
                employee.id in timesheet_map
                and str(timesheet_map[employee.id].approval_status or "").lower() == "approved"
            ),
            preserve_leave_snapshot=(
                employee.id in timesheet_map
                and str(timesheet_map[employee.id].approval_status or "").lower() == "approved"
            ),
        )
        for employee in employees
    ]

    daily_inputs = [
        FinalTimesheetDailyInput(
            employee_id=row.employee_id,
            work_date=row.work_date,
            attendance_symbol=row.attendance_symbol,
            check_in_time=row.check_in_time,
            check_out_time=row.check_out_time,
            late_minutes=int(row.late_minutes or 0),
            early_minutes=int(row.early_minutes or 0),
            abnormal_level=row.abnormal_level,
        )
        for row in daily_rows
    ]
    entry_inputs = [
        FinalTimesheetEntryInput(
            employee_id=row.employee_id,
            work_date=row.work_date,
            final_symbol=row.final_symbol,
            check_in_time=row.check_in_time,
            check_out_time=row.check_out_time,
            is_overridden=bool(row.is_overridden),
            override_reason=row.override_reason,
        )
        for row in entry_rows
    ]
    off_request_inputs = [
        FinalTimesheetOffRequestInput(
            employee_id=row.employee_id,
            request_type=row.request_type,
            start_date=row.start_date,
            end_date=row.end_date,
            total_days=float(row.total_days or 0),
            status=row.status,
        )
        for row in off_requests
    ]
    working_day_overrides = {
        row.holiday_date
        for row in db.query(HolidaySetting)
        .filter(
            HolidaySetting.is_working_day.is_(True),
            HolidaySetting.holiday_date >= period_start,
            HolidaySetting.holiday_date <= period_end,
        )
        .all()
    }

    return build_final_timesheet_report(
        period_start=period_start,
        period_end=period_end,
        employees=employee_inputs,
        daily_records=daily_inputs,
        entry_records=entry_inputs,
        off_requests=off_request_inputs,
        working_day_overrides=working_day_overrides,
    )


def build_final_timesheet_report_from_attendance_json(
    employees_payload: list[dict[str, Any]],
    period_start: date,
    period_end: date,
) -> FinalTimesheetReport:
    if not employees_payload:
        raise ValueError("Không có dữ liệu attendance JSON để export")

    employee_inputs: list[FinalTimesheetEmployeeInput] = []
    daily_inputs: list[FinalTimesheetDailyInput] = []
    export_index = 0

    for employee in employees_payload:
        attendance_details = employee.get("attendance_details") or {}
        if not isinstance(attendance_details, dict) or not attendance_details:
            continue

        export_index += 1
        machine_employee_id = str(employee.get("employee_id") or "")
        resolved_full_name = str(
            employee.get("full_name") or employee.get("employee_name") or machine_employee_id or ""
        ).strip()
        resolved_department = str(employee.get("department_name") or employee.get("department") or "") or None

        employee_inputs.append(
            FinalTimesheetEmployeeInput(
                employee_id=export_index,
                machine_employee_id=machine_employee_id,
                full_name=resolved_full_name,
                department_name=resolved_department,
                employee_type=str(employee.get("employee_type") or "FULLTIME"),
                previous_paid_leave_balance=0.0,
                current_month_paid_leave_credit=0.0,
            )
        )

        for work_date_key, raw_detail in attendance_details.items():
            try:
                work_date = date.fromisoformat(str(work_date_key))
            except ValueError:
                continue

            if not isinstance(raw_detail, dict):
                continue

            status = str(raw_detail.get("status") or "").strip()
            explicit_symbol = _clean_symbol(raw_detail.get("attendance_symbol") or raw_detail.get("final_symbol"))
            attendance_symbol = explicit_symbol or ("Ro" if status == "Absent" else None)
            abnormal_level = "L1" if status == "Missing_Punch" else None

            daily_inputs.append(
                FinalTimesheetDailyInput(
                    employee_id=export_index,
                    work_date=work_date,
                    attendance_symbol=attendance_symbol,
                    check_in_time=raw_detail.get("check_in"),
                    check_out_time=raw_detail.get("check_out"),
                    late_minutes=int(raw_detail.get("late_minutes") or 0),
                    early_minutes=0,
                    abnormal_level=abnormal_level,
                )
            )

    if not employee_inputs:
        raise ValueError("Không có dữ liệu nhân sự hợp lệ để export")

    return build_final_timesheet_report(
        period_start=period_start,
        period_end=period_end,
        employees=employee_inputs,
        daily_records=daily_inputs,
        entry_records=[],
        off_requests=[],
    )


def serialize_final_timesheet_report(report: FinalTimesheetReport) -> dict[str, Any]:
    return {
        "period_start": report.period_start.isoformat(),
        "period_end": report.period_end.isoformat(),
        "day_keys": report.day_keys,
        "day_columns": [
            {
                "key": column.key,
                "day_number": column.day_number,
                "weekday_label": column.weekday_label,
                "is_weekend": column.is_weekend,
            }
            for column in report.day_columns
        ],
        "rows": [
            {
                "employee_id": row.employee_id,
                "machine_employee_id": row.machine_employee_id,
                "full_name": row.full_name,
                "department_name": row.department_name,
                "days": row.days,
                "override_reasons": row.override_reasons,
                "abnormal_days": row.abnormal_days,
                "total_late_minutes": row.total_late_minutes,
                "total_early_minutes": row.total_early_minutes,
                "total_absent_days": row.total_absent_days,
                "total_work_days": row.total_work_days,
                "total_payroll_days": row.total_payroll_days,
                "unpaid_leave_days": row.unpaid_leave_days,
                "paid_leave_days": row.paid_leave_days,
                "previous_paid_leave_balance": row.previous_paid_leave_balance,
                "current_month_paid_leave_credit": row.current_month_paid_leave_credit,
                "remaining_paid_leave_days": row.remaining_paid_leave_days,
            }
            for row in report.rows
        ],
    }


def _format_title_period(day: date) -> str:
    return day.strftime("%d/%m/%Y")


def _display_number(value: float) -> float:
    """Keep work-unit values numeric while Excel controls their fixed display."""
    return _round_leave(value)


def export_to_final_timesheet(data: FinalTimesheetReport | dict[str, Any]) -> BytesIO:
    report = data if isinstance(data, FinalTimesheetReport) else _dict_to_report(data)
    output = BytesIO()
    summary_headers = [
        ("Nghỉ\nkhông\nlương", lambda row: row.unpaid_leave_days),
        ("Nghỉ\nhưởng\nlương", lambda row: row.paid_leave_days),
        ("Ngày phép\ncòn lại\ntháng trước", lambda row: row.previous_paid_leave_balance),
        ("Ngày phép\ntháng này", lambda row: row.current_month_paid_leave_credit),
        ("Ngày phép\ncòn lại", lambda row: row.remaining_paid_leave_days),
    ]

    def _sort_key(r):
        try:
            return float(r.machine_employee_id)
        except ValueError:
            return float('inf')

    def get_work_unit(symbol: str) -> float | str:
        if not symbol:
            return ""
        sym = _clean_symbol(symbol)
        if sym == "X": return 1.0
        if sym in ["P/Ro", "Ro/P", "X/Ro", "Ro/X"]: return 0.5
        if sym in ["X/P", "P/X", "P", "CT"]: return 1.0
        if sym in ["Ro", "0"]: return 0.0
        return ""

    sorted_rows = sorted(report.rows, key=_sort_key)
    body_rows: list[list[Any]] = []
    
    for row in sorted_rows:
        try:
            display_id = int(float(row.machine_employee_id))
        except ValueError:
            display_id = row.machine_employee_id

        ngay_cong = row.total_payroll_days
        ngay_cong_tt = row.total_work_days

        body_rows.append(
            [
                display_id,
                row.full_name,
                *[get_work_unit(row.days[key]) for key in report.day_keys],
                "", # Spacer
                _display_number(ngay_cong), # Ngay cong
                _display_number(ngay_cong_tt), # Ngay cong TT
                "", # Spacer
                *[row.days[key] for key in report.day_keys],
                *[_display_number(getter(row)) for _, getter in summary_headers],
            ]
        )

    body_df = pd.DataFrame(body_rows)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        body_df.to_excel(writer, sheet_name="Timesheet", index=False, header=False, startrow=7)
        workbook = writer.book
        sheet = workbook["Timesheet"]

        # Calculations for dimensions
        num_days = len(report.day_columns)
        num_summaries = len(summary_headers)
        total_columns = 2 + num_days + 4 + num_days + num_summaries
        
        title = f"EMPLOYEE TIMESHEET FROM {_format_title_period(report.period_start)} TO {_format_title_period(report.period_end)}"

        index_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
        weekday_fill = PatternFill(fill_type="solid", fgColor="F4B183")
        weekend_header_fill = PatternFill(fill_type="solid", fgColor="B4C6E7")
        weekend_body_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
        summary_fill = PatternFill(fill_type="solid", fgColor="C6E0B4")
        header_font = Font(bold=True)
        title_font = Font(bold=True, size=12)
        centered = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_aligned = Alignment(horizontal="left", vertical="center")
        report_border = Border(
            left=Side(style="thin", color="7F7F7F"),
            right=Side(style="thin", color="7F7F7F"),
            top=Side(style="thin", color="7F7F7F"),
            bottom=Side(style="thin", color="7F7F7F"),
        )

        title_column = 7 if total_columns >= 7 else 3
        sheet.cell(row=6, column=title_column, value=title)
        sheet.cell(row=6, column=title_column).font = title_font
        sheet.cell(row=6, column=title_column).alignment = centered

        header_row = 7
        data_start_row = 8
        sheet.cell(row=header_row, column=1, value="STT")
        sheet.cell(row=header_row, column=2, value="Họ Và Tên")
        sheet.cell(row=header_row, column=1).fill = index_fill
        sheet.cell(row=header_row, column=2).fill = index_fill

        current_col = 3
        # Block 1 headers
        for day_column in report.day_columns:
            header_value = f"{day_column.day_number}\n{day_column.weekday_label}" if day_column.is_weekend else day_column.day_number
            sheet.cell(row=header_row, column=current_col, value=header_value)
            sheet.cell(row=header_row, column=current_col).fill = weekend_header_fill if day_column.is_weekend else weekday_fill
            current_col += 1
            
        # Spacer
        current_col += 1
        
        # Ngay Cong
        sheet.cell(row=header_row, column=current_col, value="Ngày công")
        sheet.cell(row=header_row, column=current_col).fill = summary_fill
        current_col += 1
        
        # Ngay Cong TT
        sheet.cell(row=header_row, column=current_col, value="Ngày công\nTT")
        sheet.cell(row=header_row, column=current_col).fill = summary_fill
        current_col += 1
        
        # Spacer
        current_col += 1
        
        # Block 2 headers
        block_2_start = current_col
        for day_column in report.day_columns:
            header_value = f"{day_column.day_number}\n{day_column.weekday_label}" if day_column.is_weekend else day_column.day_number
            sheet.cell(row=header_row, column=current_col, value=header_value)
            sheet.cell(row=header_row, column=current_col).fill = weekend_header_fill if day_column.is_weekend else weekday_fill
            current_col += 1

        summary_start_column = current_col
        for header, _ in summary_headers:
            sheet.cell(row=header_row, column=current_col, value=header)
            sheet.cell(row=header_row, column=current_col).fill = summary_fill
            current_col += 1

        last_row = max(header_row, data_start_row + len(report.rows) - 1)
        for row in sheet.iter_rows(min_row=header_row, max_row=last_row, min_col=1, max_col=total_columns):
            for cell in row:
                cell.border = report_border
                if cell.row == header_row:
                    cell.font = header_font
                    cell.alignment = centered
                elif cell.column == 2:
                    cell.alignment = left_aligned
                else:
                    cell.alignment = centered

        # Body fills for weekends
        col_offset = 3
        for day_column in report.day_columns:
            if day_column.is_weekend:
                for row_index in range(data_start_row, last_row + 1):
                    sheet.cell(row=row_index, column=col_offset).fill = weekend_body_fill
            col_offset += 1
            
        col_offset += 4 # skip spacers and ngay cong and ngay cong tt
        
        for day_column in report.day_columns:
            if day_column.is_weekend:
                for row_index in range(data_start_row, last_row + 1):
                    sheet.cell(row=row_index, column=col_offset).fill = weekend_body_fill
            col_offset += 1

        # Summary body fill
        for column_index in range(summary_start_column, summary_start_column + len(summary_headers)):
            for row_index in range(data_start_row, last_row + 1):
                sheet.cell(row=row_index, column=column_index).fill = summary_fill

        # Work units and all calculated leave summaries must consistently show
        # one decimal place in Excel (1.0, 0.5, 0.0) instead of General's 1.
        work_unit_columns = list(range(3, 3 + num_days))
        totals_columns = [3 + num_days + 1, 3 + num_days + 2]
        numeric_columns = work_unit_columns + totals_columns + list(
            range(summary_start_column, summary_start_column + len(summary_headers))
        )
        for row_index in range(data_start_row, last_row + 1):
            for column_index in numeric_columns:
                cell = sheet.cell(row=row_index, column=column_index)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.0"

        sheet.row_dimensions[6].height = 24
        sheet.row_dimensions[7].height = 42

        sheet.column_dimensions["A"].width = 8
        sheet.column_dimensions["B"].width = 24
        
        current_col = 3
        for _ in report.day_columns:
            sheet.column_dimensions[get_column_letter(current_col)].width = 5
            current_col += 1
            
        current_col += 1 # spacer
        sheet.column_dimensions[get_column_letter(current_col)].width = 10 # Ngay cong
        current_col += 1 # Ngay cong TT
        sheet.column_dimensions[get_column_letter(current_col)].width = 10 # Ngay cong TT
        current_col += 1 # spacer
        
        for _ in report.day_columns:
            sheet.column_dimensions[get_column_letter(current_col)].width = 5
            current_col += 1

        for column_index in range(summary_start_column, summary_start_column + len(summary_headers)):
            sheet.column_dimensions[get_column_letter(column_index)].width = 12

        sheet.freeze_panes = None

    output.seek(0)
    return output

def _dict_to_report(payload: dict[str, Any]) -> FinalTimesheetReport:
    return FinalTimesheetReport(
        period_start=date.fromisoformat(payload["period_start"]),
        period_end=date.fromisoformat(payload["period_end"]),
        day_keys=list(payload.get("day_keys", [])),
        day_columns=[
            FinalTimesheetDayColumn(
                key=item["key"],
                day_number=int(item["day_number"]),
                weekday_label=str(item["weekday_label"]),
                is_weekend=bool(item["is_weekend"]),
            )
            for item in payload.get("day_columns", [])
        ],
        rows=[
            FinalTimesheetRow(
                employee_id=int(item["employee_id"]),
                machine_employee_id=str(item.get("machine_employee_id", "")),
                full_name=str(item.get("full_name", "")),
                department_name=item.get("department_name"),
                days={str(key): str(value or "") for key, value in item.get("days", {}).items()},
                override_reasons={
                    str(key): str(value or "")
                    for key, value in item.get("override_reasons", {}).items()
                },
                abnormal_days=int(item.get("abnormal_days", 0)),
                total_late_minutes=int(item.get("total_late_minutes", 0)),
                total_early_minutes=int(item.get("total_early_minutes", 0)),
                total_absent_days=_round_leave(item.get("total_absent_days", 0)),
                total_work_days=_round_leave(item.get("total_work_days", 0)),
                total_payroll_days=_round_leave(
                    item.get(
                        "total_payroll_days",
                        _round_leave(item.get("total_work_days", 0))
                        + _round_leave(item.get("paid_leave_days", 0)),
                    )
                ),
                unpaid_leave_days=_round_leave(item.get("unpaid_leave_days", 0)),
                paid_leave_days=_round_leave(item.get("paid_leave_days", 0)),
                previous_paid_leave_balance=_round_leave(item.get("previous_paid_leave_balance", 0)),
                current_month_paid_leave_credit=_round_leave(item.get("current_month_paid_leave_credit", 0)),
                remaining_paid_leave_days=_round_leave(item.get("remaining_paid_leave_days", 0)),
            )
            for item in payload.get("rows", [])
        ],
    )
