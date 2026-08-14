from io import BytesIO
import os
from copy import copy
import unicodedata
from datetime import datetime, date as date_type
from typing import Optional
import json
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models.employee import Employee
from app.models.department import Department
from app.models.monthly_salary_input import MonthlySalaryInput
from app.models.salary_policy import SalaryPolicy
from app.services.salary_policy import ensure_default_salary_policy, policy_to_dict, resolve_salary_policy


def resolve_export_salary_policy(db: Session, period: str) -> dict:
    """Use the policy snapshot already attached to the payroll month.

    This prevents a later regulation update from silently changing an already
    prepared/exported month.  Months without inputs use the version effective
    for their month.
    """
    snapshot_id = (
        db.query(MonthlySalaryInput.salary_policy_id)
        .filter(
            MonthlySalaryInput.salary_period == period,
            MonthlySalaryInput.salary_policy_id.is_not(None),
        )
        .order_by(MonthlySalaryInput.id.asc())
        .scalar()
    )
    policy = db.get(SalaryPolicy, snapshot_id) if snapshot_id else None
    if policy:
        return policy_to_dict(policy)

    # Older periods can predate the salary-policy table.  Once published they
    # are immutable, so they must remain on the baseline policy rather than
    # silently resolving to a newer regulation with a later effective date.
    is_legacy_published = db.query(MonthlySalaryInput.id).filter(
        MonthlySalaryInput.salary_period == period,
        MonthlySalaryInput.is_published.is_(True),
        MonthlySalaryInput.salary_policy_id.is_(None),
    ).first()
    if is_legacy_published:
        return policy_to_dict(ensure_default_salary_policy(db))
    return policy_to_dict(resolve_salary_policy(db, period))


def calculate_period_working_days(period_str: str) -> int:
    if not period_str:
        return 26
    try:
        from datetime import date, timedelta
        year, month = map(int, period_str.split("-"))
        prev_year = year
        prev_month = month - 1
        if prev_month == 0:
            prev_month = 12
            prev_year -= 1
        
        start_date = date(prev_year, prev_month, 23)
        end_date = date(year, month, 22)
        
        working_days = 0
        curr = start_date
        while curr <= end_date:
            if curr.weekday() < 5:  # Mon-Fri
                working_days += 1
            curr += timedelta(days=1)
        return working_days
    except Exception:
        return 26


def cake_salary(employee: dict, salary_policy: Optional[dict] = None) -> dict:
    """
    Calculates salary, deductions, taxes and transfer amount for an employee
    based on the 05/2026 Sealink salary spreadsheet specifications.
    """
    # Policy is intentionally optional for backwards-compatible exports.  New
    # payroll flows pass the effective version selected for the salary month.
    policy = salary_policy or employee.get("salary_policy") or {}
    DEDUCT_SELF = int(policy.get("personal_deduction", 15500000) or 0)
    DEDUCT_DEP = int(policy.get("dependent_deduction", 6200000) or 0)
    STANDARD_DAYS = employee.get("standard_working_days", 26)
    if not STANDARD_DAYS or STANDARD_DAYS <= 0:
        STANDARD_DAYS = 26

    contract_salary = employee.get("contract_salary", 0)
    actual_working_days = employee.get("actual_working_days", 0)
    
    # Allowances & Bonus
    meal_allowance_free = employee.get("meal_allowance_free", 0)
    meal_allowance_tax = employee.get("taxable_meal", employee.get("meal_allowance_tax", 0))
    phone_allowance_free = employee.get("phone_allowance_free", 0)
    trans_allowance_tax = employee.get("taxable_transport", employee.get("trans_allowance_tax", 0))
    performance_allowance = employee.get("performance_allowance", employee.get("perf_allowance_tax", 0))
    other_allowance = employee.get("other_allowance", employee.get("other_income", 0))
    bonus = employee.get("bonus", 0)
    bonus_14 = employee.get("bonus_14", 0)
    
    dependents_count = employee.get("dependents_count", 0)
    other_deductions = employee.get("other_deductions", 0)
    pit_refund = employee.get("pit_refund", 0)
    advance_payment = employee.get("advance_payment", 0)
    emp_type = employee.get("type", "FULLTIME")

    # 1. Mức lương thực trong tháng / Actual Salary (Col 7) & Pro-rated Allowances
    if actual_working_days <= 0:
        actual_salary = 0
        meal_allowance_free = 0
        meal_allowance_tax = 0
        phone_allowance_free = 0
        trans_allowance_tax = 0
    else:
        ratio = min(1.0, actual_working_days / STANDARD_DAYS)
        actual_salary = contract_salary if actual_working_days >= STANDARD_DAYS else round((contract_salary / STANDARD_DAYS) * actual_working_days)
        meal_allowance_free = round(meal_allowance_free * ratio)
        meal_allowance_tax = round(meal_allowance_tax * ratio)
        phone_allowance_free = round(phone_allowance_free * ratio)
        trans_allowance_tax = round(trans_allowance_tax * ratio)

    # 2. Tổng thu nhập chịu thuế TNCN / Total Taxable Income (Col 26)
    taxable_income = (
        actual_salary
        + meal_allowance_tax
        + trans_allowance_tax
        + performance_allowance
        + other_allowance
        + bonus
        + bonus_14
    )

    social_emp = 0
    health_emp = 0
    unemp_emp = 0
    total_ins_emp = 0
    
    social_comp = 0
    health_comp = 0
    unemp_comp = 0
    union_fund_comp = 0
    total_ins_comp = 0
    
    assessable_income = 0
    pit_tax = 0
    union_fee = 0
    ins_salary = 0

    if emp_type in {"PROBATION", "INTERN"} or actual_working_days <= 0:
        # Khối B Thử việc hoặc Nhân viên chưa tính lương (ngày công = 0)
        # Bảo hiểm = 0
        assessable_income = taxable_income
        if emp_type in {"PROBATION", "INTERN"} and actual_salary > 0:
            threshold = int(policy.get("probation_withholding_threshold", 2000000) or 0)
            rate = float(policy.get("probation_withholding_rate", 0.1) or 0)
            pit_tax = round(taxable_income * rate) if taxable_income >= threshold else 0
        else:
            pit_tax = 0
        union_fee = 0
        ins_salary = 0
    else:
        # Khối A Chính thức (Lương đóng BHXH là không giới hạn)
        ins_salary = contract_salary
        social_health_cap = int(policy.get("social_health_salary_cap", 0) or 0)
        if social_health_cap > 0:
            ins_salary = min(ins_salary, social_health_cap)
        
        # Employee insurance
        social_emp = round(ins_salary * float(policy.get("social_employee_rate", 0.08) or 0))
        health_emp = round(ins_salary * float(policy.get("health_employee_rate", 0.015) or 0))
        
        # special BHTN capping for employee (calculated on contract_salary + other_allowance)
        sum_contract_other = contract_salary + other_allowance
        region = str(policy.get("default_region", "I") or "I").lower()
        regional_minimum = int(policy.get(f"regional_minimum_wage_{region}", policy.get("regional_minimum_wage_i", 5310000)) or 0)
        unemployment_cap = regional_minimum * int(policy.get("unemployment_cap_multiplier", 20) or 20)
        unemp_base = min(sum_contract_other, unemployment_cap) if unemployment_cap > 0 else sum_contract_other
        unemp_emp = round(unemp_base * float(policy.get("unemployment_employee_rate", 0.01) or 0))
            
        total_ins_emp = social_emp + health_emp + unemp_emp
        
        # Employer insurance (17.5%, 3%, 1%) - Không bao gồm 2% kinh phí công đoàn trong BH bắt buộc
        social_comp = round(ins_salary * float(policy.get("social_employer_rate", 0.175) or 0))
        health_comp = round(ins_salary * float(policy.get("health_employer_rate", 0.03) or 0))
        
        # special BHTN capping for employer
        unemp_comp = round(unemp_base * float(policy.get("unemployment_employer_rate", 0.01) or 0))
            
        total_ins_comp = social_comp + health_comp + unemp_comp
        
        # Kinh phí công đoàn 2%
        union_fund_comp = round(ins_salary * float(policy.get("union_fund_employer_rate", 0.02) or 0))

        assessable_income = max(
            0,
            taxable_income
            - total_ins_emp
            - DEDUCT_SELF
            - (dependents_count * DEDUCT_DEP),
        )

        ai = assessable_income
        default_brackets = [
            {"up_to": 10000000, "rate": 0.05, "deduction": 0},
            {"up_to": 30000000, "rate": 0.10, "deduction": 500000},
            {"up_to": 60000000, "rate": 0.20, "deduction": 3500000},
            {"up_to": 100000000, "rate": 0.30, "deduction": 9500000},
            {"up_to": None, "rate": 0.35, "deduction": 14500000},
        ]
        brackets = policy.get("pit_brackets")
        if not brackets and policy.get("pit_brackets_json"):
            try:
                brackets = json.loads(policy["pit_brackets_json"])
            except (json.JSONDecodeError, TypeError):
                brackets = None
        brackets = brackets or default_brackets
        selected_bracket = brackets[-1]
        for bracket in brackets:
            ceiling = bracket.get("up_to")
            if ceiling is None or ai <= float(ceiling):
                selected_bracket = bracket
                break
        pit = ai * float(selected_bracket.get("rate", 0)) - float(selected_bracket.get("deduction", 0))
        pit_tax = round(max(0.0, pit))
        
        # Union fee: Khối chính thức trích 0.5% tính trên Lương nộp BHXH, tối đa 10% lương tối thiểu chung (234,000 VND)
        union_cap = int(policy.get("union_employee_cap", 234000) or 0)
        union_fee = round(min(ins_salary * float(policy.get("union_employee_rate", 0.005) or 0), union_cap))

    # 3. Lương thực nhận (NET)
    # Col 29: NET Salary = ROUND(actual_salary + total_allowances_all - total_ins_emp - pit_tax, 2)
    total_allowances_all = (
        meal_allowance_free
        + meal_allowance_tax
        + phone_allowance_free
        + trans_allowance_tax
        + performance_allowance
        + other_allowance
        + bonus
        + bonus_14
    )
    
    net_salary = round(
        actual_salary
        + total_allowances_all
        - total_ins_emp
        - pit_tax,
        2
    )

    # 4. Thực chuyển qua ngân hàng (Phải là số DƯƠNG)
    # Col 33: Total Transfer = net_salary + pit_refund - union_fee - other_deductions
    total_transfer = max(0, round(net_salary + pit_refund - union_fee - other_deductions, 2))
    # Col 35: Rest Transfer = total_transfer - advance_payment
    final_transfer = max(0, round(total_transfer - advance_payment, 2))

    return {
        "actual_salary": actual_salary,
        "meal_allowance_free": meal_allowance_free,
        "meal_allowance_tax": meal_allowance_tax,
        "phone_allowance_free": phone_allowance_free,
        "trans_allowance_tax": trans_allowance_tax,
        "taxable_income": taxable_income,
        "assessable_income": assessable_income,
        "ins_salary": ins_salary,
        "social_emp": social_emp,
        "health_emp": health_emp,
        "unemp_emp": unemp_emp,
        "total_ins_emp": total_ins_emp,
        "social_comp": social_comp,
        "health_comp": health_comp,
        "unemp_comp": unemp_comp,
        "union_fund_comp": union_fund_comp,
        "total_ins_comp": total_ins_comp,
        "pit_tax": pit_tax,
        "union_fee": union_fee,
        "net_salary": net_salary,
        "total_transfer": total_transfer,
        "final_transfer": final_transfer,
    }


def remove_vietnamese_accents(s: str) -> str:
    if not s:
        return ""
    normalized = unicodedata.normalize('NFKD', s)
    cleaned = ''.join([c for c in normalized if not unicodedata.combining(c)])
    return cleaned.replace('Đ', 'D').replace('đ', 'd').replace('đ'.upper(), 'D')


def get_monthly_input_fields(emp, input_rec, period: str = None) -> dict:
    """
    Extracts the monthly salary fields with standard defaults from input_rec,
    falling back to values from the employee record or default system values
    to match the frontend grid values.
    """
    fullname = input_rec.fullname or emp.full_name if input_rec else emp.full_name
    position = input_rec.position or emp.position if input_rec else emp.position
    contract_salary = input_rec.contract_salary if input_rec and input_rec.contract_salary is not None else emp.contract_salary
    dependents_count = input_rec.dependents_count if input_rec and input_rec.dependents_count is not None else emp.dependents_count

    std_days = calculate_period_working_days(period)
    actual_days = input_rec.actual_working_days if (input_rec and input_rec.actual_working_days is not None) else float(std_days)
    meal_free = input_rec.meal_allowance_free if (input_rec and input_rec.meal_allowance_free is not None) else emp.meal_allowance
    meal_tax = input_rec.meal_allowance_tax if (input_rec and input_rec.meal_allowance_tax is not None) else 0
    phone_free = input_rec.phone_allowance_free if (input_rec and input_rec.phone_allowance_free is not None) else emp.phone_allowance
    trans_tax = input_rec.trans_allowance_tax if (input_rec and input_rec.trans_allowance_tax is not None) else emp.trans_allowance
    perf_tax = input_rec.perf_allowance_tax if (input_rec and input_rec.perf_allowance_tax is not None) else emp.other_allowance
    other_inc = input_rec.other_income if (input_rec and input_rec.other_income is not None) else 0
    bonus = input_rec.bonus if (input_rec and input_rec.bonus is not None) else 0
    bonus_14 = input_rec.bonus_14 if (input_rec and input_rec.bonus_14 is not None) else 0
    pit_refund = input_rec.pit_refund if (input_rec and input_rec.pit_refund is not None) else 0
    other_ded = input_rec.other_deductions if (input_rec and input_rec.other_deductions is not None) else 0
    advance = input_rec.advance_payment if (input_rec and input_rec.advance_payment is not None) else 0

    return {
        "fullname": fullname,
        "position": position,
        "contract_salary": contract_salary,
        "dependents_count": dependents_count,
        "actual_days": actual_days,
        "meal_free": meal_free,
        "meal_tax": meal_tax,
        "phone_free": phone_free,
        "trans_tax": trans_tax,
        "perf_tax": perf_tax,
        "other_inc": other_inc,
        "bonus": bonus,
        "bonus_14": bonus_14,
        "pit_refund": pit_refund,
        "other_ded": other_ded,
        "advance": advance,
    }


def get_row_styles(ws, row_idx):
    styles = {}
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col)
        styles[col] = {
            "font": copy(cell.font) if cell.font else None,
            "border": copy(cell.border) if cell.border else None,
            "fill": copy(cell.fill) if cell.fill else None,
            "alignment": copy(cell.alignment) if cell.alignment else None,
            "number_format": cell.number_format,
        }
    return styles


def apply_row_styles(ws, row_idx, styles):
    for col, style in styles.items():
        cell = ws.cell(row=row_idx, column=col)
        if style["font"]:
            cell.font = style["font"]
        if style["border"]:
            cell.border = style["border"]
        if style["fill"]:
            cell.fill = style["fill"]
        if style["alignment"]:
            cell.alignment = style["alignment"]
        if style["number_format"]:
            cell.number_format = style["number_format"]


def export_salary_report(db: Session, period: str) -> BytesIO:
    effective_policy = resolve_export_salary_policy(db, period)
    # 1. Fetch employees and monthly inputs
    results = db.query(Employee, MonthlySalaryInput).outerjoin(
        MonthlySalaryInput,
        (MonthlySalaryInput.employee_id == Employee.id) & (MonthlySalaryInput.salary_period == period)
    ).order_by(Employee.id.asc()).all()

    # Python filtering
    filtered_results = []
    for emp, m_input in results:
        # If there's an input record, always keep it
        if m_input is not None:
            filtered_results.append((emp, m_input))
            continue
            
        # Filter by start date
        if emp.start_date:
            start_period = emp.start_date.strftime("%Y-%m")
            if start_period > period:
                continue
                
        # Filter by resignation period
        if emp.status == 'RESIGNED':
            if emp.resignation_period:
                if emp.resignation_period <= period:
                    continue
            else:
                continue
        elif emp.status == 'LOCKED':
            continue

        filtered_results.append((emp, m_input))

    # 2. Partition
    def get_resolved_type(emp, m_input):
        if m_input and m_input.employee_type is not None:
            return m_input.employee_type
        return emp.employee_type

    fulltime_emps = [pair for pair in filtered_results if get_resolved_type(pair[0], pair[1]) == "FULLTIME"]
    probation_emps = [
        pair for pair in filtered_results
        if get_resolved_type(pair[0], pair[1]) in {"PROBATION", "INTERN"}
    ]

    # Load template relative to this file's location
    template_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "resources", "salary_template.xlsx")
    wb = openpyxl.load_workbook(template_path)
    ws_salary = wb['Employee salary']
    ws_bank = wb['Bank Transfer']
    ws_sealink = wb['Sealink Transfer']

    # --- 3. Process Sheet: Employee salary ---
    # Store styles of key rows
    style_fulltime_data = get_row_styles(ws_salary, 11)
    style_subtotal_a = get_row_styles(ws_salary, 13)
    style_probation_header = get_row_styles(ws_salary, 14)
    style_probation_data = get_row_styles(ws_salary, 15)
    style_subtotal_b = get_row_styles(ws_salary, 16)
    style_grand_total = get_row_styles(ws_salary, 18)
    
    # Footer styles (rows 20 to 30)
    footer_styles = {}
    for r in range(20, 31):
        footer_styles[r] = {
            "styles": get_row_styles(ws_salary, r),
            "height": ws_salary.row_dimensions[r].height,
            "values": [ws_salary.cell(row=r, column=c).value for c in range(1, ws_salary.max_column + 1)]
        }

    # Clear merged cell ranges below row 9
    merged_ranges_to_remove = []
    for r_range in list(ws_salary.merged_cells.ranges):
        if r_range.bounds[1] > 9: # start_row > 9
            merged_ranges_to_remove.append(r_range)
    for r_range in merged_ranges_to_remove:
        ws_salary.merged_cells.remove(r_range)

    # Clean rows starting from row 10 to the bottom
    ws_salary.delete_rows(10, ws_salary.max_row - 9 + 5)

    year, month = period.split("-")
    
    # Write Group A Header
    ws_salary.cell(row=10, column=1, value="A. NHÂN VIÊN CÓ HỢP ĐỒNG LAO ĐỘNG/ EMPLOYEE WITH LABOR CONTRACT")
    ws_salary.merge_cells(start_row=10, start_column=1, end_row=10, end_column=27)
    apply_row_styles(ws_salary, 10, style_fulltime_data) # use fulltime cell format for styling
    ws_salary.cell(row=10, column=1).font = openpyxl.styles.Font(name="Times New Roman", size=15, bold=True)
    
    row_cursor = 11
    added_row_numbers_a = []
    
    # Write Group A Data
    std_days = calculate_period_working_days(period)
    for stt, (emp, input_rec) in enumerate(fulltime_emps, 1):
        fields = get_monthly_input_fields(emp, input_rec, period)
        sales_bonus = round(get_sales_bonus_for_employee_period(db, emp.id, period), 2)
        
        # Prepare dict for calculation (using old logic)
        emp_data = {
            "type": "FULLTIME",
            "contract_salary": fields["contract_salary"],
            "actual_working_days": fields["actual_days"],
            "standard_working_days": std_days,
            "meal_allowance_free": fields["meal_free"],
            "meal_allowance_tax": fields["meal_tax"],
            "phone_allowance_free": fields["phone_free"],
            "taxable_transport": fields["trans_tax"],
            "performance_allowance": fields["perf_tax"],
            "other_allowance": fields["other_inc"],
            "bonus": fields["bonus"] + sales_bonus,
            "bonus_14": fields["bonus_14"],
            "dependents_count": fields["dependents_count"],
            "other_deductions": fields["other_ded"],
            "pit_refund": fields["pit_refund"],
            "advance_payment": fields["advance"],
        }
        res = cake_salary(emp_data, effective_policy)
        
        # Write exact values calculated by Python backend
        ws_salary.cell(row=row_cursor, column=1, value=stt)
        ws_salary.cell(row=row_cursor, column=2, value=fields["actual_days"])
        ws_salary.cell(row=row_cursor, column=3, value=fields["fullname"])
        ws_salary.cell(row=row_cursor, column=4, value=fields["dependents_count"])
        ws_salary.cell(row=row_cursor, column=5, value=fields["position"] or "")
        ws_salary.cell(row=row_cursor, column=6, value=fields["contract_salary"])
        ws_salary.cell(row=row_cursor, column=7, value=res["actual_salary"])
        ws_salary.cell(row=row_cursor, column=8, value=res["meal_allowance_free"])
        ws_salary.cell(row=row_cursor, column=9, value=fields["meal_tax"])
        ws_salary.cell(row=row_cursor, column=10, value=res["phone_allowance_free"])
        ws_salary.cell(row=row_cursor, column=11, value=res["trans_allowance_tax"])
        ws_salary.cell(row=row_cursor, column=12, value=fields["perf_tax"])
        ws_salary.cell(row=row_cursor, column=13, value=fields["other_inc"])
        ws_salary.cell(row=row_cursor, column=14, value=fields["bonus"] + sales_bonus)
        ws_salary.cell(row=row_cursor, column=15, value=fields["bonus_14"])
        ws_salary.cell(row=row_cursor, column=16, value=res["ins_salary"])
        ws_salary.cell(row=row_cursor, column=17, value=res["social_emp"])
        ws_salary.cell(row=row_cursor, column=18, value=res["health_emp"])
        ws_salary.cell(row=row_cursor, column=19, value=res["unemp_emp"])
        ws_salary.cell(row=row_cursor, column=20, value=res["total_ins_emp"])
        ws_salary.cell(row=row_cursor, column=21, value=res["social_comp"])
        ws_salary.cell(row=row_cursor, column=22, value=res["health_comp"])
        ws_salary.cell(row=row_cursor, column=23, value=res["unemp_comp"])
        ws_salary.cell(row=row_cursor, column=24, value=res["total_ins_comp"])
        ws_salary.cell(row=row_cursor, column=25, value=res["union_fund_comp"])
        ws_salary.cell(row=row_cursor, column=26, value=res["taxable_income"])
        ws_salary.cell(row=row_cursor, column=27, value=res["assessable_income"])
        ws_salary.cell(row=row_cursor, column=28, value=res["pit_tax"])
        ws_salary.cell(row=row_cursor, column=29, value=res["net_salary"])
        ws_salary.cell(row=row_cursor, column=30, value=res["union_fee"])
        ws_salary.cell(row=row_cursor, column=31, value=fields["other_ded"])
        ws_salary.cell(row=row_cursor, column=32, value=fields["pit_refund"])
        ws_salary.cell(row=row_cursor, column=33, value=res["total_transfer"])
        ws_salary.cell(row=row_cursor, column=34, value=fields["advance"])
        ws_salary.cell(row=row_cursor, column=35, value=res["final_transfer"])
        
        apply_row_styles(ws_salary, row_cursor, style_fulltime_data)
        
        # Set blue text color for name
        ws_salary.cell(row=row_cursor, column=3).font = openpyxl.styles.Font(name="Times New Roman", size=15, color="FF0070C0")
        
        added_row_numbers_a.append(row_cursor)
        row_cursor += 1

    # Write Subtotal Group A
    r_sub_a = row_cursor
    ws_salary.cell(row=r_sub_a, column=1, value="Sub-total")
    ws_salary.merge_cells(start_row=r_sub_a, start_column=1, end_row=r_sub_a, end_column=3)
    apply_row_styles(ws_salary, r_sub_a, style_subtotal_a)
    
    # Populate SUM formulas for subtotal A
    for c in range(4, 36):
        if c in [5]: # position is text
            continue
        col_letter = get_column_letter(c)
        if added_row_numbers_a:
            ws_salary.cell(row=r_sub_a, column=c, value=f"=SUM({col_letter}{added_row_numbers_a[0]}:{col_letter}{added_row_numbers_a[-1]})")
        else:
            ws_salary.cell(row=r_sub_a, column=c, value=0)
            
    row_cursor += 1
    
    # Write Group B Title Row
    r_title_b = row_cursor
    ws_salary.cell(row=r_title_b, column=1, value="B. NHÂN VIÊN THỬ VIỆC + HỌC VIỆC / PROBATIONARY+ APPRENTICE STAFF - THUẾ 10%")
    ws_salary.merge_cells(start_row=r_title_b, start_column=1, end_row=r_title_b, end_column=27)
    ws_salary.cell(row=r_title_b, column=28, value="Thuế suất\nTax rate")
    ws_salary.cell(row=r_title_b, column=29, value=0.1)
    apply_row_styles(ws_salary, r_title_b, style_probation_header)
    
    row_cursor += 1
    added_row_numbers_b = []
    
    # Write Group B Data
    std_days = calculate_period_working_days(period)
    for stt_b, (emp, input_rec) in enumerate(probation_emps, 1):
        fields = get_monthly_input_fields(emp, input_rec, period)
        sales_bonus = round(get_sales_bonus_for_employee_period(db, emp.id, period), 2)
        
        # Prepare dict for calculation (using old logic)
        emp_data = {
            "type": "PROBATION",
            "contract_salary": fields["contract_salary"],
            "actual_working_days": fields["actual_days"],
            "standard_working_days": std_days,
            "meal_allowance_free": fields["meal_free"],
            "meal_allowance_tax": fields["meal_tax"],
            "phone_allowance_free": fields["phone_free"],
            "taxable_transport": fields["trans_tax"],
            "performance_allowance": fields["perf_tax"],
            "other_allowance": fields["other_inc"],
            "bonus": fields["bonus"] + sales_bonus,
            "bonus_14": fields["bonus_14"],
            "dependents_count": fields["dependents_count"],
            "other_deductions": fields["other_ded"],
            "pit_refund": fields["pit_refund"],
            "advance_payment": fields["advance"],
        }
        res = cake_salary(emp_data, effective_policy)
        
        # Write exact values calculated by Python backend
        ws_salary.cell(row=row_cursor, column=1, value=stt_b)
        ws_salary.cell(row=row_cursor, column=2, value=fields["actual_days"])
        ws_salary.cell(row=row_cursor, column=3, value=fields["fullname"])
        ws_salary.cell(row=row_cursor, column=4, value=0)
        ws_salary.cell(row=row_cursor, column=5, value=fields["position"] or "")
        ws_salary.cell(row=row_cursor, column=6, value=fields["contract_salary"])
        ws_salary.cell(row=row_cursor, column=7, value=res["actual_salary"])
        ws_salary.cell(row=row_cursor, column=8, value=res["meal_allowance_free"])
        ws_salary.cell(row=row_cursor, column=9, value=fields["meal_tax"])
        ws_salary.cell(row=row_cursor, column=10, value=res["phone_allowance_free"])
        ws_salary.cell(row=row_cursor, column=11, value=res["trans_allowance_tax"])
        ws_salary.cell(row=row_cursor, column=12, value=fields["perf_tax"])
        ws_salary.cell(row=row_cursor, column=13, value=fields["other_inc"])
        ws_salary.cell(row=row_cursor, column=14, value=fields["bonus"] + sales_bonus)
        ws_salary.cell(row=row_cursor, column=15, value=fields["bonus_14"])
        ws_salary.cell(row=row_cursor, column=16, value=res["ins_salary"])
        ws_salary.cell(row=row_cursor, column=17, value=res["social_emp"])
        ws_salary.cell(row=row_cursor, column=18, value=res["health_emp"])
        ws_salary.cell(row=row_cursor, column=19, value=res["unemp_emp"])
        ws_salary.cell(row=row_cursor, column=20, value=res["total_ins_emp"])
        ws_salary.cell(row=row_cursor, column=21, value=res["social_comp"])
        ws_salary.cell(row=row_cursor, column=22, value=res["health_comp"])
        ws_salary.cell(row=row_cursor, column=23, value=res["unemp_comp"])
        ws_salary.cell(row=row_cursor, column=24, value=res["total_ins_comp"])
        ws_salary.cell(row=row_cursor, column=25, value=res["union_fund_comp"])
        ws_salary.cell(row=row_cursor, column=26, value=res["taxable_income"])
        ws_salary.cell(row=row_cursor, column=27, value=res["assessable_income"])
        ws_salary.cell(row=row_cursor, column=28, value=res["pit_tax"])
        ws_salary.cell(row=row_cursor, column=29, value=res["net_salary"])
        ws_salary.cell(row=row_cursor, column=30, value=res["union_fee"])
        ws_salary.cell(row=row_cursor, column=31, value=fields["other_ded"])
        ws_salary.cell(row=row_cursor, column=32, value=fields["pit_refund"])
        ws_salary.cell(row=row_cursor, column=33, value=res["total_transfer"])
        ws_salary.cell(row=row_cursor, column=34, value=fields["advance"])
        ws_salary.cell(row=row_cursor, column=35, value=res["final_transfer"])
        
        apply_row_styles(ws_salary, row_cursor, style_probation_data)
        added_row_numbers_b.append(row_cursor)
        row_cursor += 1

    # Write Subtotal Group B
    r_sub_b = row_cursor
    ws_salary.cell(row=r_sub_b, column=1, value="Sub-total")
    ws_salary.merge_cells(start_row=r_sub_b, start_column=1, end_row=r_sub_b, end_column=3)
    apply_row_styles(ws_salary, r_sub_b, style_subtotal_b)
    
    # Populate SUM formulas for subtotal B
    for c in range(4, 36):
        if c in [4, 5]: # dependents and position
            continue
        col_letter = get_column_letter(c)
        if added_row_numbers_b:
            ws_salary.cell(row=r_sub_b, column=c, value=f"=SUM({col_letter}{added_row_numbers_b[0]}:{col_letter}{added_row_numbers_b[-1]})")
        else:
            ws_salary.cell(row=r_sub_b, column=c, value=0)
            
    row_cursor += 1
    
    # Empty Row
    row_cursor += 1
    
    # Write Grand Total Row
    r_total = row_cursor
    ws_salary.cell(row=r_total, column=1, value="TỔNG CỘNG")
    ws_salary.merge_cells(start_row=r_total, start_column=1, end_row=r_total, end_column=3)
    apply_row_styles(ws_salary, r_total, style_grand_total)
    
    for c in range(4, 36):
        if c in [5]:
            continue
        col_letter = get_column_letter(c)
        ws_salary.cell(row=r_total, column=c, value=f"={col_letter}{r_sub_a}+{col_letter}{r_sub_b}")
        
    row_cursor += 2
    
    # Footer and signatures offset row calculation
    shift = row_cursor - 20
    
    # Write Footer
    for orig_r in sorted(footer_styles.keys()):
        target_r = orig_r + shift
        ws_salary.row_dimensions[target_r].height = footer_styles[orig_r]["height"]
        
        # Apply style of original row
        apply_row_styles(ws_salary, target_r, footer_styles[orig_r]["styles"])
        
        # Write values and formulas
        vals = footer_styles[orig_r]["values"]
        for c_idx, val in enumerate(vals, 1):
            if val is not None:
                # Update any formulas in footer to point to shifted rows
                if isinstance(val, str) and val.startswith("="):
                    # Replace totals and subtotals references
                    val = val.replace("T18", f"T{r_total}")
                    val = val.replace("AB18", f"AB{r_total}")
                    val = val.replace("AC18", f"AC{r_total}")
                    val = val.replace("AG18", f"AG{r_total}")
                    val = val.replace("AI18", f"AI{r_total}")
                    val = val.replace("AI13", f"AI{r_sub_a}")
                    val = val.replace("AI16", f"AI{r_sub_b}")
                    val = val.replace("G18", f"G{r_total}")
                    val = val.replace("H18", f"H{r_total}")
                    val = val.replace("I18", f"I{r_total}")
                    val = val.replace("J18", f"J{r_total}")
                    val = val.replace("K18", f"K{r_total}")
                    val = val.replace("L18", f"L{r_total}")
                    val = val.replace("M18", f"M{r_total}")
                    val = val.replace("N18", f"N{r_total}")
                    val = val.replace("AD18", f"AD{r_total}")
                    
                    # Row-local reference logic in footer row 21 and 23
                    if orig_r == 21:
                        val = val.replace("AI21", f"AI{21+shift}")
                        val = val.replace("AC21", f"AC{21+shift}")
                    elif orig_r == 23:
                        val = val.replace("AC21", f"AC{21+shift}")
                        val = val.replace("N22", f"N{22+shift}")
                    elif orig_r == 27:
                        val = val.replace("AG21", f"AG{21+shift}")
                        
                ws_salary.cell(row=target_r, column=c_idx, value=val)
                
    # Re-apply merged cell ranges in footer
    ws_salary.merge_cells(start_row=20+shift, start_column=17, end_row=20+shift, end_column=27) # date
    ws_salary.merge_cells(start_row=21+shift, start_column=17, end_row=21+shift, end_column=27) # phê duyệt
    ws_salary.merge_cells(start_row=22+shift, start_column=17, end_row=22+shift, end_column=27) # approved by
    ws_salary.merge_cells(start_row=27+shift, start_column=21, end_row=27+shift, end_column=23) # ton that trung kien

    # --- 4. Process Sheet: Bank Transfer ---
    # Store styles
    style_bank_data = get_row_styles(ws_bank, 6)
    style_bank_total = get_row_styles(ws_bank, 8)
    style_bank_sig = get_row_styles(ws_bank, 10)
    style_bank_prepared = get_row_styles(ws_bank, 11)
    style_bank_director = get_row_styles(ws_bank, 16)
    
    # Store heights
    height_bank_data = ws_bank.row_dimensions[6].height
    height_bank_total = ws_bank.row_dimensions[8].height
    height_bank_sig = ws_bank.row_dimensions[10].height
    height_bank_prepared = ws_bank.row_dimensions[11].height
    height_bank_director = ws_bank.row_dimensions[16].height

    # Clear merged ranges below row 5
    merged_ranges_bank_to_remove = []
    for r_range in list(ws_bank.merged_cells.ranges):
        if r_range.bounds[1] > 5:
            merged_ranges_bank_to_remove.append(r_range)
    for r_range in merged_ranges_bank_to_remove:
        ws_bank.merged_cells.remove(r_range)
        
    # Clear all data starting from row 6
    ws_bank.delete_rows(6, ws_bank.max_row - 5 + 5)

    all_active_emps = fulltime_emps + probation_emps
    bank_cursor = 6
    
    std_days = calculate_period_working_days(period)
    for stt_bk, (emp, input_rec) in enumerate(all_active_emps, 1):
        fullname = input_rec.fullname or emp.full_name if input_rec else emp.full_name
        account_number = input_rec.account_number or emp.account_number if input_rec else emp.account_number
        bank_name = input_rec.bank_name or emp.bank_name if input_rec else emp.bank_name
        
        fields = get_monthly_input_fields(emp, input_rec, period)
        emp_type = get_resolved_type(emp, input_rec)
        sales_bonus = round(get_sales_bonus_for_employee_period(db, emp.id, period), 2)
        
        emp_data = {
            "type": emp_type,
            "contract_salary": fields["contract_salary"],
            "actual_working_days": fields["actual_days"],
            "standard_working_days": std_days,
            "meal_allowance_free": fields["meal_free"],
            "meal_allowance_tax": fields["meal_tax"],
            "phone_allowance_free": fields["phone_free"],
            "taxable_transport": fields["trans_tax"],
            "performance_allowance": fields["perf_tax"],
            "other_allowance": fields["other_inc"],
            "bonus": fields["bonus"] + sales_bonus,
            "bonus_14": fields["bonus_14"],
            "dependents_count": fields["dependents_count"],
            "other_deductions": fields["other_ded"],
            "pit_refund": fields["pit_refund"],
            "advance_payment": fields["advance"],
        }
        res = cake_salary(emp_data, effective_policy)
        
        ws_bank.row_dimensions[bank_cursor].height = height_bank_data
        ws_bank.cell(row=bank_cursor, column=1, value=stt_bk).alignment = openpyxl.styles.Alignment(horizontal="center")
        ws_bank.cell(row=bank_cursor, column=2, value=fullname).alignment = openpyxl.styles.Alignment(horizontal="left")
        ws_bank.cell(row=bank_cursor, column=3, value=res["final_transfer"]).alignment = openpyxl.styles.Alignment(horizontal="right")
        ws_bank.cell(row=bank_cursor, column=4, value=account_number or "").alignment = openpyxl.styles.Alignment(horizontal="center")
        ws_bank.cell(row=bank_cursor, column=5, value=bank_name or "").alignment = openpyxl.styles.Alignment(horizontal="left")
        ws_bank.cell(row=bank_cursor, column=6, value="HCM").alignment = openpyxl.styles.Alignment(horizontal="center")
        ws_bank.cell(row=bank_cursor, column=7, value="Lương tháng").alignment = openpyxl.styles.Alignment(horizontal="left")
        ws_bank.cell(row=bank_cursor, column=8, value="=$E$3").alignment = openpyxl.styles.Alignment(horizontal="left")
        
        apply_row_styles(ws_bank, bank_cursor, style_bank_data)
        ws_bank.cell(row=bank_cursor, column=3).font = openpyxl.styles.Font(name="Times New Roman", size=12, color="FF00B0F0")
        
        bank_cursor += 1

    # Write Total row in Bank Transfer
    r_bank_total = bank_cursor
    ws_bank.row_dimensions[r_bank_total].height = height_bank_total
    ws_bank.cell(row=r_bank_total, column=1, value="")
    ws_bank.cell(row=r_bank_total, column=2, value="")
    ws_bank.cell(row=r_bank_total, column=3, value=f"=SUM(C6:C{r_bank_total-1})")
    ws_bank.cell(row=r_bank_total, column=4, value="")
    ws_bank.cell(row=r_bank_total, column=5, value=f"='Employee salary'!$Q${20+shift}")
    
    apply_row_styles(ws_bank, r_bank_total, style_bank_total)
    ws_bank.merge_cells(start_row=r_bank_total, start_column=5, end_row=r_bank_total, end_column=8)
    
    bank_cursor += 1
    
    # Write Signatures in Bank Transfer
    r_sig_h = bank_cursor + 1
    ws_bank.row_dimensions[r_sig_h].height = height_bank_sig
    ws_bank.cell(row=r_sig_h, column=2, value="LẬP BIỂU")
    ws_bank.cell(row=r_sig_h, column=4, value="KIỂM TRA")
    ws_bank.cell(row=r_sig_h, column=6, value="PHÊ DUYỆT")
    apply_row_styles(ws_bank, r_sig_h, style_bank_sig)
    
    ws_bank.merge_cells(start_row=r_sig_h, start_column=2, end_row=r_sig_h, end_column=3)
    ws_bank.merge_cells(start_row=r_sig_h, start_column=4, end_row=r_sig_h, end_column=5)
    ws_bank.merge_cells(start_row=r_sig_h, start_column=6, end_row=r_sig_h, end_column=8)
    
    r_sig_sub = r_sig_h + 1
    ws_bank.row_dimensions[r_sig_sub].height = height_bank_prepared
    ws_bank.cell(row=r_sig_sub, column=2, value="PREPARED BY")
    ws_bank.cell(row=r_sig_sub, column=4, value="CHECKED BY")
    ws_bank.cell(row=r_sig_sub, column=6, value="APPROVED BY")
    apply_row_styles(ws_bank, r_sig_sub, style_bank_prepared)
    
    ws_bank.merge_cells(start_row=r_sig_sub, start_column=2, end_row=r_sig_sub, end_column=3)
    ws_bank.merge_cells(start_row=r_sig_sub, start_column=4, end_row=r_sig_sub, end_column=5)
    ws_bank.merge_cells(start_row=r_sig_sub, start_column=6, end_row=r_sig_sub, end_column=8)
    
    r_sig_name = r_sig_sub + 5
    ws_bank.row_dimensions[r_sig_name].height = height_bank_director
    ws_bank.cell(row=r_sig_name, column=6, value="TÔN THẤT TRUNG KIÊN")
    apply_row_styles(ws_bank, r_sig_name, style_bank_director)
    ws_bank.merge_cells(start_row=r_sig_name, start_column=6, end_row=r_sig_name, end_column=8)

    # --- 5. Process Sheet: Sealink Transfer ---
    style_sealink_data = get_row_styles(ws_sealink, 2)
    height_sealink_data = ws_sealink.row_dimensions[2].height
    
    # Delete sample data rows starting from row 2
    ws_sealink.delete_rows(2, ws_sealink.max_row - 1 + 5)
    
    sealink_cursor = 2
    for stt_sl, (emp, input_rec) in enumerate(all_active_emps, 1):
        fullname = input_rec.fullname or emp.full_name if input_rec else emp.full_name
        account_number = input_rec.account_number or emp.account_number if input_rec else emp.account_number
        bank_name = input_rec.bank_name or emp.bank_name if input_rec else emp.bank_name
        
        fields = get_monthly_input_fields(emp, input_rec, period)
        emp_type = get_resolved_type(emp, input_rec)
        sales_bonus = round(get_sales_bonus_for_employee_period(db, emp.id, period), 2)
        
        emp_data = {
            "type": emp_type,
            "contract_salary": fields["contract_salary"],
            "actual_working_days": fields["actual_days"],
            "standard_working_days": std_days,
            "meal_allowance_free": fields["meal_free"],
            "meal_allowance_tax": fields["meal_tax"],
            "phone_allowance_free": fields["phone_free"],
            "taxable_transport": fields["trans_tax"],
            "performance_allowance": fields["perf_tax"],
            "other_allowance": fields["other_inc"],
            "bonus": fields["bonus"] + sales_bonus,
            "bonus_14": fields["bonus_14"],
            "dependents_count": fields["dependents_count"],
            "other_deductions": fields["other_ded"],
            "pit_refund": fields["pit_refund"],
            "advance_payment": fields["advance"],
        }
        res = cake_salary(emp_data, effective_policy)
        
        clean_name = remove_vietnamese_accents(fullname).upper()
        clean_bank = remove_vietnamese_accents(bank_name).upper() if bank_name else ""
        
        ws_sealink.row_dimensions[sealink_cursor].height = height_sealink_data
        ws_sealink.cell(row=sealink_cursor, column=1, value=stt_sl).alignment = openpyxl.styles.Alignment(horizontal="center")
        ws_sealink.cell(row=sealink_cursor, column=2, value="")
        ws_sealink.cell(row=sealink_cursor, column=3, value=account_number or "").alignment = openpyxl.styles.Alignment(horizontal="center")
        ws_sealink.cell(row=sealink_cursor, column=4, value="")
        ws_sealink.cell(row=sealink_cursor, column=5, value="")
        ws_sealink.cell(row=sealink_cursor, column=6, value="")
        ws_sealink.cell(row=sealink_cursor, column=7, value=clean_name).alignment = openpyxl.styles.Alignment(horizontal="left")
        ws_sealink.cell(row=sealink_cursor, column=8, value=clean_bank).alignment = openpyxl.styles.Alignment(horizontal="left")
        ws_sealink.cell(row=sealink_cursor, column=9, value=res["final_transfer"]).alignment = openpyxl.styles.Alignment(horizontal="right")
        ws_sealink.cell(row=sealink_cursor, column=10, value="vnd").alignment = openpyxl.styles.Alignment(horizontal="center")
        ws_sealink.cell(row=sealink_cursor, column=11, value=f"payroll {month}-{year[2:]}").alignment = openpyxl.styles.Alignment(horizontal="left")
        
        apply_row_styles(ws_sealink, sealink_cursor, style_sealink_data)
        sealink_cursor += 1

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


from app.models.department_bonus_config import DepartmentBonusConfig

DEFAULT_BONUS_RULES = [
    {"min": 0, "max": 2.0, "rate": 0.0},
    {"min": 2.01, "max": 4.0, "rate": 0.20},
    {"min": 4.01, "max": 6.0, "rate": 0.25},
    {"min": 6.01, "max": 8.0, "rate": 0.30},
    {"min": 8.01, "max": 999.0, "rate": 0.35}
]

FIXED_NON_SALES_BONUS_RATE = 0.20
FIXED_NON_SALES_BONUS_RULES = [
    {"min": 0, "max": 999.0, "rate": FIXED_NON_SALES_BONUS_RATE}
]
SALES_BONUS_DEPARTMENT_NAMES = {"SALE LOCAL", "SALE OVERSEA"}


def _normalize_bonus_department_name(value: Optional[str]) -> str:
    return " ".join(str(value or "").upper().replace("_", " ").replace("-", " ").split())


def is_sales_bonus_department(*department_values: Optional[str]) -> bool:
    return any(
        _normalize_bonus_department_name(value) in SALES_BONUS_DEPARTMENT_NAMES
        for value in department_values
        if value
    )


def is_sales_bonus_employee(employee: Employee) -> bool:
    department = getattr(employee, "department", None)
    department_name = getattr(department, "name", None)
    if department_name:
        return is_sales_bonus_department(department_name)
    return is_sales_bonus_department(
        getattr(employee, "department_name", None),
        getattr(employee, "department_code", None),
    )

def get_active_department_rules(db: Session, department_id: Optional[int], period: str) -> list:
    if not department_id:
        return [dict(rule) for rule in DEFAULT_BONUS_RULES]

    department = db.query(Department).filter(Department.id == department_id).first()
    if department and not is_sales_bonus_department(department.name):
        return [dict(rule) for rule in FIXED_NON_SALES_BONUS_RULES]

    configs = (
        db.query(DepartmentBonusConfig)
        .filter(
            DepartmentBonusConfig.department_id == department_id,
            DepartmentBonusConfig.period <= period
        )
        .order_by(DepartmentBonusConfig.period.desc())
        .all()
    )
    for config in configs:
        if config.end_period is None or config.end_period >= period:
            return config.rules
            
    return [dict(rule) for rule in DEFAULT_BONUS_RULES]

def calculateDynamicSalesBonus(gross_profit: float, employee_salary: float, rules: list = None) -> dict:
    """
    Calculates Sales Bonus with dynamic progressive tiers based on Target:
    - Bước 1: Tính Net Profit = Gross_Profit * 0.95
    - Bước 2: Tính Target = Employee_Salary * base_coef
    - Bước 3: Tính phần lũy tiến ban đầu PF_COUNT_BN = Net_Profit - Target
      (Nếu <= 0, Bonus = 0)
    - Bước 4: Lũy tiến phần PF_COUNT_BN theo các mốc từ rules
    """
    if rules is None:
        rules = DEFAULT_BONUS_RULES

    net_profit = gross_profit * 0.95
    if employee_salary <= 0:
        return {
            "net_profit": round(net_profit, 2),
            "target": 0.0,
            "pf_count_bn": 0.0,
            "profit_count_bonus": 0.0,
            "bonus_rate": 0.0,
            "total_bonus_quarter": 0.0,
            "bonus_per_month": 0.0,
            "coefficient": 0.0
        }
    
    sorted_rules = sorted(rules, key=lambda x: x["min"])
    base_rule = sorted_rules[0]
    base_coef = base_rule["max"]
    
    target = employee_salary * base_coef
    pf_count_bn = net_profit - target
    
    coef = round(net_profit / employee_salary, 2)
    
    total_bonus_quarter = 0.0
    effective_rate = 0.0
    
    if pf_count_bn > 0:
        remaining = pf_count_bn
        prev_max = base_coef
        
        for rule in sorted_rules[1:]:
            rate = rule["rate"]
            rule_max = rule["max"]
            
            if rule_max >= 999.0:
                if remaining > 0:
                    total_bonus_quarter += remaining * rate
                    remaining = 0
            else:
                tier_coef_size = rule_max - prev_max
                if tier_coef_size > 0 and remaining > 0:
                    tier_profit_size = tier_coef_size * employee_salary
                    amount = min(remaining, tier_profit_size)
                    total_bonus_quarter += amount * rate
                    remaining -= amount
                prev_max = rule_max
                
        effective_rate = (total_bonus_quarter / pf_count_bn) if pf_count_bn > 0 else 0.0
        
    bonus_per_month = total_bonus_quarter / 3.0
    
    return {
        "net_profit": round(net_profit, 2),
        "target": round(target, 2),
        "pf_count_bn": round(pf_count_bn, 2) if pf_count_bn > 0 else 0.0,
        "profit_count_bonus": round(pf_count_bn, 2) if pf_count_bn > 0 else 0.0,
        "bonus_rate": round(effective_rate, 4),
        "total_bonus_quarter": round(total_bonus_quarter, 2),
        "bonus_per_month": round(bonus_per_month, 2),
        "coefficient": coef
    }


def calculate_employee_bonus(
    gross_profit: float,
    employee_salary: float,
    rules: list = None,
    uses_progressive_bonus: bool = True,
) -> dict:
    """Keep the existing SALE calculation and apply one fixed tier elsewhere."""
    if uses_progressive_bonus:
        return calculateDynamicSalesBonus(gross_profit, employee_salary, rules)

    net_profit = gross_profit * 0.95
    eligible_profit = max(net_profit, 0.0)
    total_bonus_quarter = eligible_profit * FIXED_NON_SALES_BONUS_RATE
    coefficient = round(net_profit / employee_salary, 2) if employee_salary > 0 else 0.0
    return {
        "net_profit": round(net_profit, 2),
        "target": 0.0,
        "pf_count_bn": round(eligible_profit, 2),
        "profit_count_bonus": round(eligible_profit, 2),
        "bonus_rate": FIXED_NON_SALES_BONUS_RATE,
        "total_bonus_quarter": round(total_bonus_quarter, 2),
        "bonus_per_month": round(total_bonus_quarter / 3.0, 2),
        "coefficient": coefficient,
    }


def calculate_sales_bonus(gross_profit: float, target: float = 120000000.0) -> dict:
    """
    Wrapper for backward compatibility. Computes employee_salary as target / 2.
    """
    return calculateDynamicSalesBonus(gross_profit, employee_salary=target / 2.0)


def clean_name_for_match(name: str) -> str:
    if not name:
        return ""
    import unicodedata
    normalized = unicodedata.normalize('NFKD', name)
    cleaned = ''.join([c for c in normalized if not unicodedata.combining(c)])
    return "".join(cleaned.lower().replace('đ', 'd').replace('Đ', 'D').split())


def parse_date_string(s: str) -> Optional[date_type]:
    s = s.strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%b-%Y", "%d-%B-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    try:
        parts = s.replace("/", "-").replace(" ", "-").split("-")
        if len(parts) == 3:
            day = int(parts[0])
            month_str = parts[1].lower()
            year = int(parts[2])
            months = {
                "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
                "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
                "thang1": 1, "thang2": 2, "thang3": 3, "thang4": 4, "thang5": 5,
                "thang6": 6, "thang7": 7, "thang8": 8, "thang9": 9, "thang10": 10,
                "thang11": 11, "thang12": 12
            }
            month_full = {
                "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
                "july": 7, "august": 8, "september": 9, "october": 10, "november": 11, "december": 12
            }
            m = months.get(month_str[:3]) or month_full.get(month_str)
            if not m and month_str.isdigit():
                m = int(month_str)
            if m:
                return date_type(year, m, day)
    except Exception:
        pass
    return None


def get_period_dates(period) -> tuple:
    if period.from_date and period.till_date:
        return period.from_date, period.till_date
    
    label = str(period.period_label or "")
    import re
    dates = re.findall(r"\d{2}[-\/\s](?:[a-zA-Z]{3}|\d{2})[-\/\s]\d{4}", label)
    if len(dates) >= 2:
        from_d = parse_date_string(dates[0])
        till_d = parse_date_string(dates[1])
        if from_d and till_d:
            return from_d, till_d
    return None, None


def get_sales_bonus_for_employee_period(db: Session, employee_id: int, period: str) -> float:
    """
    Calculates the sales bonus for an employee for a specific payout period (month YYYY-MM).
    Payout month mapping to source quarter:
    - Month 1, 2, 3 (Q1) -> Source: Q4 of previous year (months 10, 11, 12)
    - Month 4, 5, 6 (Q2) -> Source: Q1 of current year (months 1, 2, 3)
    - Month 7, 8, 9 (Q3) -> Source: Q2 of current year (months 4, 5, 6)
    - Month 10, 11, 12 (Q4) -> Source: Q3 of current year (months 7, 8, 9)
    """
    employee = db.query(Employee).filter(Employee.id == employee_id).first()
    if not employee:
        return 0.0
        
    # Once a commission wallet exists, it is the financial source of truth.
    # The old calculation below intentionally remains as a fallback only for
    # historical commission imports that have not been synced to the wallet.
    # This makes holds, manual adjustments, transfers and payout schedules
    # immediately visible in the salary grid without touching the base salary
    # calculation.
    wallet_bonus = get_wallet_sales_bonus_for_employee_period(db, employee, period)
    if wallet_bonus is not None:
        return wallet_bonus

    uses_progressive_bonus = is_sales_bonus_employee(employee)

    # SALE keeps the existing enablement rule. Non-SALE uses the fixed formula
    # regardless of its old department bonus configuration.
    if uses_progressive_bonus and (
        employee.bonus_coefficient is None or float(employee.bonus_coefficient) <= 0.0
    ):
        return 0.0

    try:
        year = int(period[:4])
        month = int(period[5:7])
    except (ValueError, TypeError, IndexError):
        return 0.0
        
    # Determine the source quarter date range based on payout month
    if month in [1, 2, 3]:
        source_year = year - 1
        source_start = date_type(source_year, 10, 1)
        source_end = date_type(source_year, 12, 31)
    elif month in [4, 5, 6]:
        source_start = date_type(year, 1, 1)
        source_end = date_type(year, 3, 31)
    elif month in [7, 8, 9]:
        source_start = date_type(year, 4, 1)
        source_end = date_type(year, 6, 30)
    else:
        source_start = date_type(year, 7, 1)
        source_end = date_type(year, 9, 30)
        
    from app.models.commission import CommissionPeriod, CommissionJob
    
    # Get all commission periods that overlap with the source quarter range
    periods = db.query(CommissionPeriod).all()
    matching_period_ids = []
    for p in periods:
        fd, td = get_period_dates(p)
        if fd and td:
            if source_start <= fd <= source_end:
                matching_period_ids.append(p.id)
                
    if not matching_period_ids:
        return 0.0
        
    # Match the employee to sales_rep names in CommissionJob and overrides
    emp_clean = clean_name_for_match(employee.full_name)
    
    from app.models.commission import CommissionRepOverride
    
    total_sales_bonus = 0.0
    
    for p_id in matching_period_ids:
        # Check if there is an override for this employee in this period
        overrides = db.query(CommissionRepOverride).filter(
            CommissionRepOverride.period_id == p_id
        ).all()
        emp_ov = None
        for o in overrides:
            if clean_name_for_match(o.sales_rep) == emp_clean:
                emp_ov = o
                break
        
        # Calculate original gross profit from jobs of this period
        jobs_of_p = db.query(CommissionJob).filter(
            CommissionJob.period_id == p_id
        ).all()
        rep_jobs = [j for j in jobs_of_p if clean_name_for_match(j.sales_rep) == emp_clean]
        gross_profit = sum(j.profit_loss for j in rep_jobs)
        
        period_bonus = 0.0
        
        if emp_ov:
            # If override exists, apply override values where present
            if emp_ov.override_monthly_bonus is not None:
                period_bonus = emp_ov.override_monthly_bonus
            elif emp_ov.override_total_bonus is not None:
                period_bonus = emp_ov.override_total_bonus / 3.0
            else:
                # Need to recalculate based on pnl, target, and rate overrides
                pnl = emp_ov.override_profit_loss if emp_ov.override_profit_loss is not None else gross_profit
                if not uses_progressive_bonus:
                    bonus_rate = (
                        emp_ov.override_bonus_rate
                        if emp_ov.override_bonus_rate is not None
                        else FIXED_NON_SALES_BONUS_RATE
                    )
                    period_bonus = (max(pnl * 0.95, 0.0) * bonus_rate) / 3.0
                else:
                    net_profit = pnl * 0.95
                    target = emp_ov.override_target if emp_ov.override_target is not None else (float(employee.contract_salary or 0.0) * 2.0)
                    pf_count_bn = net_profit - target
                    if pf_count_bn <= 0:
                        period_bonus = 0.0
                    else:
                        if emp_ov.override_bonus_rate is not None:
                            bonus_rate = emp_ov.override_bonus_rate
                        else:
                            salary = float(employee.contract_salary or 0.0)
                            if salary <= 0.0:
                                salary = target / 2.0
                            if salary > 0.0:
                                coef = round(net_profit / salary, 2)
                                bonus_dept_id = int(employee.bonus_coefficient) if employee.bonus_coefficient and float(employee.bonus_coefficient) > 0 else employee.department_id
                                rules = get_active_department_rules(db, bonus_dept_id, period)
                                bonus_rate = 0.0
                                for rule in sorted(rules, key=lambda x: x["min"]):
                                    if rule["min"] <= coef <= rule["max"]:
                                        bonus_rate = rule["rate"]
                                        break
                            else:
                                bonus_rate = 0.0
                        period_bonus = (pf_count_bn * bonus_rate) / 3.0
        else:
            # Fallback to normal calculations from jobs
            if gross_profit > 0:
                bonus_dept_id = int(employee.bonus_coefficient) if employee.bonus_coefficient and float(employee.bonus_coefficient) > 0 else employee.department_id
                rules = get_active_department_rules(db, bonus_dept_id, period)
                bonus_info = calculate_employee_bonus(
                    gross_profit,
                    employee_salary=float(employee.contract_salary or 0.0),
                    rules=rules,
                    uses_progressive_bonus=uses_progressive_bonus,
                )
                period_bonus = bonus_info.get("bonus_per_month", 0.0)
                
        total_sales_bonus += period_bonus

    return total_sales_bonus


def _wallet_payout_months(commission_period) -> list[str]:
    """Return the three monthly payroll periods following a source quarter."""
    _, till_date = get_period_dates(commission_period)
    if not till_date:
        return []
    months: list[str] = []
    year, month = till_date.year, till_date.month
    for _ in range(3):
        year += 1 if month == 12 else 0
        month = 1 if month == 12 else month + 1
        months.append(f"{year:04d}-{month:02d}")
    return months


def _wallet_position_amounts(entries) -> dict:
    """Mirror the wallet's ledger-derived available balance locally.

    This function deliberately derives a value from immutable ledger rows;
    it does not persist or modify any commission amount.
    """
    earned_types = {
        "ACCRUAL_HELD", "ACCRUAL_AVAILABLE", "ADJUSTMENT_HELD",
        "ADJUSTMENT_AVAILABLE", "REVERSAL_HELD", "REVERSAL_AVAILABLE",
        "REVERSAL_PAID", "MANUAL_CREDIT", "MANUAL_DECREASE",
    }
    payment_hold_types = {
        "ACCRUAL_HELD", "ADJUSTMENT_HELD", "REVERSAL_HELD",
        "PAYMENT_STATUS_HOLD",
    }
    calculation_earned = payment_held = released = 0.0
    for entry in entries:
        entry_type = str(entry.entry_type or "")
        amount = float(entry.amount or 0.0)
        if entry_type in earned_types and entry_type not in {"MANUAL_CREDIT", "MANUAL_DECREASE"}:
            calculation_earned += amount
        if entry_type in payment_hold_types:
            payment_held += amount
        elif entry_type == "RELEASED":
            released += amount
    automatic_held = max(0.0, round(payment_held - released, 2))
    return {
        "calculation_earned": round(calculation_earned, 2),
        "automatic_held": automatic_held,
    }


def get_wallet_sales_bonus_for_employee_period(db: Session, employee: Employee, period: str) -> Optional[float]:
    """Return the monthly salary commission derived from the wallet.

    A commission quarter creates *three* monthly entitlement amounts.  Its
    calculated monthly amount is therefore the base for every following
    payroll month (for example Q2 -> July, August and September).  Ledger
    adjustments are applied only to their selected payroll month.  This avoids
    the old error where the whole three-month bonus was placed in July.
    """
    from app.models.commission import CommissionCalculationSnapshot, CommissionPeriod, CommissionWalletLedger, CommissionPayoutSchedule

    emp_name = clean_name_for_match(employee.full_name)
    all_entries = db.query(CommissionWalletLedger).order_by(CommissionWalletLedger.id.asc()).all()
    entries = [
        entry for entry in all_entries
        if entry.employee_id == employee.id or clean_name_for_match(entry.sales_rep) == emp_name
    ]
    if not entries:
        return None

    # A JOB payment command is different from a normal payout schedule: it
    # moves a previously-held amount from its source quarter to a future
    # commission cycle. Only these schedules are added to their target payroll
    # month; generic schedules merely reserve an already-available amount.
    active_command_schedule_ids = {
        row.id for row in db.query(CommissionPayoutSchedule.id).filter(
            CommissionPayoutSchedule.payment_verification_id.is_not(None),
            CommissionPayoutSchedule.status.in_({"SCHEDULED", "PAID"}),
        ).all()
    }

    positions: dict[tuple[int, Optional[int], str], list] = {}
    for entry in entries:
        positions.setdefault((entry.period_id, entry.job_id, entry.sales_rep), []).append(entry)

    positions_by_period: dict[int, list[list]] = {}
    for (period_id, _job_id, _sales_rep), position_entries in positions.items():
        positions_by_period.setdefault(period_id, []).append(position_entries)

    result = 0.0
    has_wallet_data_for_month = False
    month_adjustment_types = {
        "MANUAL_CREDIT", "MANUAL_DECREASE", "MANUAL_HOLD",
        "MANUAL_RELEASE", "TRANSFER_OUT",
    }
    for period_id, period_positions in positions_by_period.items():
        source_period = db.get(CommissionPeriod, period_id)
        payout_months = _wallet_payout_months(source_period) if source_period else []
        if period not in payout_months:
            continue
        has_wallet_data_for_month = True
        period_position = {
            "calculation_earned": round(sum(_wallet_position_amounts(item)["calculation_earned"] for item in period_positions), 2),
            "automatic_held": round(sum(_wallet_position_amounts(item)["automatic_held"] for item in period_positions), 2),
        }
        payout_index = payout_months.index(period)
        snapshot = db.query(CommissionCalculationSnapshot).filter(
            CommissionCalculationSnapshot.period_id == period_id,
        ).filter(
            (CommissionCalculationSnapshot.employee_id == employee.id) |
            (CommissionCalculationSnapshot.sales_rep == employee.full_name),
        ).order_by(CommissionCalculationSnapshot.id.desc()).first()
        if snapshot and float(snapshot.total_bonus_quarter or 0.0) > 0:
            # Round the first two months, then place the fractional remainder
            # in the final month. This guarantees that three monthly rows add
            # up exactly to the saved quarterly commission.
            quarter_total = round(float(snapshot.total_bonus_quarter), 2)
            regular_month = round(quarter_total / 3, 2)
            monthly_base = regular_month if payout_index < 2 else round(quarter_total - regular_month * 2, 2)
        else:
            monthly_base = period_position["calculation_earned"]
        # The calculated amount is a monthly entitlement. A Payment Received
        # hold belongs to the source quarter and is distributed precisely over
        # its three payroll months. When it has a future release allocation,
        # use the original held source amount; otherwise use the current
        # unpaid hold balance (important for legacy records and YES -> NO).
        result += monthly_base
        source_hold = 0.0
        for position_entries in period_positions:
            amounts = _wallet_position_amounts(position_entries)
            release_allocated = sum(
                float(entry.amount or 0.0)
                for entry in position_entries
                if entry.entry_type in {"PAYMENT_RELEASE_ALLOCATION", "PAYMENT_RELEASE_REVERSAL"}
            )
            command_scheduled = sum(
                float(entry.amount or 0.0)
                for entry in position_entries
                if entry.entry_type == "SCHEDULED" and entry.schedule_id in active_command_schedule_ids
            )
            # Current unpaid hold plus the still-active future release plan
            # equals the one source hold that must be spread across this
            # quarter. This remains correct after YES -> NO -> YES changes.
            source_hold += amounts["automatic_held"] + max(0.0, round(release_allocated, 2)) + max(0.0, round(command_scheduled, 2))
        source_hold = round(source_hold, 2)
        hold_parts = [round(source_hold / 3, 2), round(source_hold / 3, 2)]
        hold_parts.append(round(source_hold - hold_parts[0] - hold_parts[1], 2))
        result -= hold_parts[payout_index]
        default_adjustment_month = payout_months[0]
        for entry in (entry for position_entries in period_positions for entry in position_entries):
            if entry.entry_type not in month_adjustment_types:
                continue
            adjustment_month = entry.payout_period or default_adjustment_month
            if adjustment_month != period:
                continue
            if entry.entry_type == "MANUAL_HOLD":
                result -= abs(float(entry.amount or 0.0))
            elif entry.entry_type == "MANUAL_RELEASE":
                result += abs(float(entry.amount or 0.0))
            else:
                # MANUAL_DECREASE and TRANSFER_OUT are persisted as negative
                # amounts; a manual credit is positive.
                result += float(entry.amount or 0.0)

    # Transfer in is the one adjustment that may target a month outside the
    # normal three-month range of its source commission quarter.
    transfer_in = sum(
        float(entry.amount or 0.0)
        for entry in entries
        if entry.entry_type == "TRANSFER_IN" and entry.payout_period == period
    )
    if transfer_in:
        has_wallet_data_for_month = True
        result += transfer_in
    payment_release = sum(
        float(entry.amount or 0.0)
        for entry in entries
        if entry.entry_type in {"PAYMENT_RELEASE_ALLOCATION", "PAYMENT_RELEASE_REVERSAL"}
        and entry.payout_period == period
    )
    if payment_release:
        has_wallet_data_for_month = True
        result += payment_release
    command_payout = sum(
        float(entry.amount or 0.0)
        for entry in entries
        if entry.entry_type == "SCHEDULED"
        and entry.schedule_id in active_command_schedule_ids
        and entry.payout_period == period
    )
    if command_payout:
        has_wallet_data_for_month = True
        result += command_payout
    return round(result, 2) if has_wallet_data_for_month else None


def _wallet_source_bonus_for_month(
    db: Session,
    employee: Employee,
    source_period,
    period_positions: list[list],
    payout_period: str,
) -> float:
    """Calculate one source-quarter contribution for one payroll month.

    This is a read-only view of the same ledger rules used by
    ``get_wallet_sales_bonus_for_employee_period``.  It lets the employee
    payslip explain a quarter without changing any payroll/commission rule.
    """
    from app.models.commission import CommissionCalculationSnapshot

    payout_months = _wallet_payout_months(source_period)
    if payout_period not in payout_months:
        return 0.0

    payout_index = payout_months.index(payout_period)
    period_position = {
        "calculation_earned": round(sum(_wallet_position_amounts(item)["calculation_earned"] for item in period_positions), 2),
    }
    snapshot = db.query(CommissionCalculationSnapshot).filter(
        CommissionCalculationSnapshot.period_id == source_period.id,
    ).filter(
        (CommissionCalculationSnapshot.employee_id == employee.id) |
        (CommissionCalculationSnapshot.sales_rep == employee.full_name),
    ).order_by(CommissionCalculationSnapshot.id.desc()).first()
    if snapshot and float(snapshot.total_bonus_quarter or 0.0) > 0:
        quarter_total = round(float(snapshot.total_bonus_quarter), 2)
        regular_month = round(quarter_total / 3, 2)
        result = regular_month if payout_index < 2 else round(quarter_total - regular_month * 2, 2)
    else:
        result = period_position["calculation_earned"]

    source_hold = 0.0
    for position_entries in period_positions:
        amounts = _wallet_position_amounts(position_entries)
        release_allocated = sum(
            float(entry.amount or 0.0)
            for entry in position_entries
            if entry.entry_type in {"PAYMENT_RELEASE_ALLOCATION", "PAYMENT_RELEASE_REVERSAL"}
        )
        source_hold += amounts["automatic_held"] + max(0.0, round(release_allocated, 2))
    hold_parts = [round(source_hold / 3, 2), round(source_hold / 3, 2)]
    hold_parts.append(round(source_hold - hold_parts[0] - hold_parts[1], 2))
    result -= hold_parts[payout_index]

    adjustment_types = {"MANUAL_CREDIT", "MANUAL_DECREASE", "MANUAL_HOLD", "MANUAL_RELEASE", "TRANSFER_OUT"}
    default_adjustment_month = payout_months[0]
    for entry in (entry for position_entries in period_positions for entry in position_entries):
        if entry.entry_type not in adjustment_types:
            continue
        if (entry.payout_period or default_adjustment_month) != payout_period:
            continue
        amount = float(entry.amount or 0.0)
        result += -abs(amount) if entry.entry_type == "MANUAL_HOLD" else abs(amount) if entry.entry_type == "MANUAL_RELEASE" else amount
    return round(result, 2)


def get_commission_payslip_summary(db: Session, employee: Employee, payout_period: str) -> dict:
    """Return a transparent, read-only commission breakdown for a payslip.

    ``remaining_bonus`` means the planned amount in the *later months* of the
    same three-month payout cycle. Jobs with Payment Received other than YES
    are listed separately, with the ledger amount currently waiting for the
    customer payment condition.
    """
    from app.models.commission import CommissionCalculationSnapshot, CommissionJob, CommissionPeriod, CommissionPayoutSchedule, CommissionWalletLedger

    employee_name = clean_name_for_match(employee.full_name)
    entries = [
        entry for entry in db.query(CommissionWalletLedger).order_by(CommissionWalletLedger.id.asc()).all()
        if entry.employee_id == employee.id or clean_name_for_match(entry.sales_rep) == employee_name
    ]
    positions: dict[tuple[int, Optional[int], str], list] = {}
    for entry in entries:
        positions.setdefault((entry.period_id, entry.job_id, entry.sales_rep), []).append(entry)

    positions_by_period: dict[int, list[list]] = {}
    for (period_id, _job_id, _sales_rep), position_entries in positions.items():
        positions_by_period.setdefault(period_id, []).append(position_entries)

    cycles: list[dict] = []
    pending_jobs: list[dict] = []
    # Payment commands are additional bonus amounts released from a previously
    # held JOB.  Keep the accountant's note with the target payslip month so
    # the employee can see both the amount and its source without altering any
    # salary calculation.
    active_command_schedules = {
        item.id: item
        for item in db.query(CommissionPayoutSchedule).filter(
            CommissionPayoutSchedule.payment_verification_id.is_not(None),
            CommissionPayoutSchedule.status.in_({"SCHEDULED", "PAID"}),
        ).all()
    }
    scheduled_job_payouts: list[dict] = []
    for entry in entries:
        schedule = active_command_schedules.get(entry.schedule_id)
        if entry.entry_type != "SCHEDULED" or not schedule or entry.payout_period != payout_period:
            continue
        job = db.get(CommissionJob, entry.job_id) if entry.job_id else None
        source_period = db.get(CommissionPeriod, entry.period_id)
        scheduled_job_payouts.append({
            "job_no": job.job_no if job else "JOB commission",
            "customer": job.customer if job else None,
            "source_period_label": source_period.period_label if source_period else None,
            "amount": round(float(entry.amount or 0.0), 2),
            "note": entry.note or schedule.note or "Kế toán đã lập lệnh chi trả bonus theo JOB.",
        })
    for period_id, period_positions in positions_by_period.items():
        source_period = db.get(CommissionPeriod, period_id)
        if not source_period:
            continue
        payout_months = _wallet_payout_months(source_period)
        release_current = sum(
            float(entry.amount or 0.0)
            for position_entries in period_positions for entry in position_entries
            if entry.entry_type in {"PAYMENT_RELEASE_ALLOCATION", "PAYMENT_RELEASE_REVERSAL"}
            and entry.payout_period == payout_period
        )
        if payout_period not in payout_months and abs(release_current) < 0.005:
            continue

        snapshot = db.query(CommissionCalculationSnapshot).filter(
            CommissionCalculationSnapshot.period_id == period_id,
        ).filter(
            (CommissionCalculationSnapshot.employee_id == employee.id) |
            (CommissionCalculationSnapshot.sales_rep == employee.full_name),
        ).order_by(CommissionCalculationSnapshot.id.desc()).first()
        quarter_total = round(float(snapshot.total_bonus_quarter or 0.0), 2) if snapshot else 0.0
        regular_current = _wallet_source_bonus_for_month(db, employee, source_period, period_positions, payout_period)
        future_regular = sum(
            _wallet_source_bonus_for_month(db, employee, source_period, period_positions, month)
            for month in payout_months if month > payout_period
        )
        future_releases = sum(
            float(entry.amount or 0.0)
            for position_entries in period_positions for entry in position_entries
            if entry.entry_type in {"PAYMENT_RELEASE_ALLOCATION", "PAYMENT_RELEASE_REVERSAL"}
            and entry.payout_period and entry.payout_period > payout_period
        )
        cycles.append({
            "period_id": period_id,
            "period_label": source_period.period_label,
            "payout_periods": payout_months,
            "total_bonus_quarter": quarter_total,
            "current_period_bonus": round(regular_current + release_current, 2),
            "remaining_bonus": round(future_regular + future_releases, 2),
        })

        for job in db.query(CommissionJob).filter(CommissionJob.period_id == period_id).all():
            if clean_name_for_match(job.sales_rep) != employee_name or str(job.payment_received or "").strip().upper() in {"YES", "Y", "PAID", "TRUE", "1"}:
                continue
            position_entries = positions.get((period_id, job.id, job.sales_rep or ""), [])
            job_monthly_bonus = _wallet_position_amounts(position_entries)["calculation_earned"] if position_entries else 0.0
            # A zero-value JOB does not give the employee useful pending-bonus
            # information. Keep it in the immutable ledger, but omit it from
            # the personal payslip list to avoid empty rows.
            if round(job_monthly_bonus, 2) < 0.005:
                continue
            pending_jobs.append({
                "period_label": source_period.period_label,
                "job_no": job.job_no,
                "customer": job.customer,
                "payment_received": job.payment_received or "NO",
                "pending_bonus": round(job_monthly_bonus, 2),
                "reason": "Chờ khách hàng thanh toán (Payment Received = NO).",
            })

    actual_current = get_sales_bonus_for_employee_period(db, employee.id, payout_period)
    return {
        "payout_period": payout_period,
        "cycles": cycles,
        "total_bonus_quarter": round(sum(item["total_bonus_quarter"] for item in cycles), 2),
        "current_period_bonus": round(actual_current, 2),
        "remaining_bonus": round(sum(item["remaining_bonus"] for item in cycles), 2),
        "pending_jobs": pending_jobs,
        "pending_bonus_amount": round(sum(item["pending_bonus"] for item in pending_jobs), 2),
        "scheduled_job_payouts": scheduled_job_payouts,
        "scheduled_job_payout_total": round(sum(item["amount"] for item in scheduled_job_payouts), 2),
    }
