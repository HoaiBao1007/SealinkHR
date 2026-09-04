"""Parse ageing/payment reports used to reconcile commission JOB hold bonus."""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
import re
import unicodedata

from openpyxl import load_workbook


FIXED_HOLD_BONUS_PERCENT = 30.0


@dataclass(frozen=True)
class ReceivableJobBalance:
    job_no: str
    source_rows: int
    receivable_amount: float
    received_amount: float
    balance_amount: float
    currency: str

    @property
    def outstanding_percent(self) -> float:
        if self.receivable_amount <= 0:
            return 0.0
        return max(0.0, self.balance_amount / self.receivable_amount * 100)

    @property
    def paid_percent(self) -> float:
        return max(0.0, min(100.0, 100.0 - self.outstanding_percent))

    @property
    def hold_bonus_percent(self) -> float:
        """A fully-paid JOB has no hold; outstanding JOBs retain 30%."""
        return 0.0 if self.balance_amount <= 0 else FIXED_HOLD_BONUS_PERCENT


@dataclass(frozen=True)
class ReceivableParseResult:
    sheet_name: str
    header_row: int
    jobs: tuple[ReceivableJobBalance, ...]
    positive_rows: int
    ignored_non_positive_rows: int
    invalid_positive_rows: int


def normalize_receivable_job_no(value: object) -> str:
    text = str(value or "").strip().upper()
    text = text.translate(str.maketrans({"–": "-", "—": "-", "−": "-"}))
    return re.sub(r"\s+", "", text)


def _normalize_header(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(character for character in text if not unicodedata.combining(character))
    return re.sub(r"[^A-Z0-9]+", "", text.upper())


def _as_number(value: object) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("\u00a0", "").replace(" ", "")
    negative_parentheses = text.startswith("(") and text.endswith(")")
    if negative_parentheses:
        text = text[1:-1]
    text = text.replace(",", "")
    try:
        number = float(text)
    except ValueError:
        return 0.0
    return -number if negative_parentheses else number


def _find_header(worksheet) -> tuple[int, dict[str, int]] | None:
    aliases = {
        "job_no": {"JOBNO", "JOBNUMBER", "JOB"},
        "receivable": {"RECEIVABLEPAYABLE", "RECEIVABLE", "TOTALRECEIVABLE"},
        "received": {"RECEIVEDPAID", "RECEIVED", "AMOUNTRECEIVED"},
        "balance": {"BALANCE", "OUTSTANDINGBALANCE", "REMAININGBALANCE"},
        "currency": {"CUR", "CURRENCY"},
    }
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=min(30, worksheet.max_row), values_only=True),
        start=1,
    ):
        normalized = [_normalize_header(value) for value in row]
        columns: dict[str, int] = {}
        for key, candidates in aliases.items():
            index = next((position for position, value in enumerate(normalized) if value in candidates), None)
            if index is not None:
                columns[key] = index
        if {"job_no", "receivable", "balance"}.issubset(columns):
            return row_number, columns
    return None


def parse_receivable_workbook(contents: bytes, filename: str) -> ReceivableParseResult:
    suffix = Path(filename or "").suffix.lower()
    if suffix != ".xlsx":
        raise ValueError("File đối chiếu công nợ phải có định dạng .xlsx.")
    try:
        workbook = load_workbook(BytesIO(contents), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Không thể đọc file Excel công nợ hoặc file đã bị hỏng.") from exc

    try:
        selected = None
        for worksheet in workbook.worksheets:
            header = _find_header(worksheet)
            if header:
                selected = (worksheet, *header)
                break
        if selected is None:
            raise ValueError("Không tìm thấy đủ các cột Job No, Receivable / Payable và Balance.")

        worksheet, header_row, columns = selected
        aggregated: dict[str, dict[str, object]] = {}
        positive_rows = 0
        ignored_non_positive_rows = 0
        invalid_positive_rows = 0
        for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            balance = _as_number(row[columns["balance"]] if columns["balance"] < len(row) else None)
            # A negative balance is an overpayment/credit and is intentionally
            # ignored.  Balance = 0 is different: the customer has paid the
            # receivable in full, so keep the row to record Payment Received
            # and an explicit Hold Bonus of 0.
            if balance < 0:
                ignored_non_positive_rows += 1
                continue
            positive_rows += 1
            job_no = normalize_receivable_job_no(row[columns["job_no"]] if columns["job_no"] < len(row) else None)
            receivable = _as_number(row[columns["receivable"]] if columns["receivable"] < len(row) else None)
            if not job_no or receivable <= 0:
                invalid_positive_rows += 1
                continue
            received = _as_number(row[columns["received"]] if columns.get("received", -1) < len(row) and columns.get("received", -1) >= 0 else None)
            currency = str(row[columns["currency"]] if columns.get("currency", -1) < len(row) and columns.get("currency", -1) >= 0 else "VND").strip().upper() or "VND"
            current = aggregated.setdefault(job_no, {
                "rows": 0,
                "receivable": 0.0,
                "received": 0.0,
                "balance": 0.0,
                "currency": currency,
            })
            if current["currency"] != currency:
                invalid_positive_rows += 1
                continue
            current["rows"] = int(current["rows"]) + 1
            current["receivable"] = float(current["receivable"]) + receivable
            current["received"] = float(current["received"]) + max(0.0, received)
            current["balance"] = float(current["balance"]) + balance

        jobs = tuple(
            ReceivableJobBalance(
                job_no=job_no,
                source_rows=int(values["rows"]),
                receivable_amount=round(float(values["receivable"]), 2),
                received_amount=round(float(values["received"]), 2),
                balance_amount=round(float(values["balance"]), 2),
                currency=str(values["currency"]),
            )
            for job_no, values in aggregated.items()
        )
        if not jobs:
            raise ValueError("File không có dòng Balance bằng 0 hoặc dương hợp lệ để đối chiếu.")
        return ReceivableParseResult(
            sheet_name=worksheet.title,
            header_row=header_row,
            jobs=jobs,
            positive_rows=positive_rows,
            ignored_non_positive_rows=ignored_non_positive_rows,
            invalid_positive_rows=invalid_positive_rows,
        )
    finally:
        workbook.close()
