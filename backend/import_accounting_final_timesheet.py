r"""Import an accountant-approved attendance workbook into the website.

Both ``Timesheet`` and ``TTS`` are supported. Calendar working-day overrides
are scoped to individual dates, so the normal weekend rule stays unchanged.

Usage (from backend):
    .venv\Scripts\python.exe import_accounting_final_timesheet.py "<file.xlsx>" \
        --working-day 2026-08-22 \
        --working-day-reason "Làm bù lễ 02/09/2026"
"""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from app.db.session import SessionLocal
from app.models.attendance_daily import AttendanceDaily
from app.models.employee import Employee
from app.models.holiday_setting import HolidaySetting
from app.models.monthly_salary_input import MonthlySalaryInput
from app.models.timesheet import Timesheet
from app.models.timesheet_entry import TimesheetEntry
from app.models.timesheet_period import TimesheetPeriod
from app.services.final_timesheet_report import _clean_symbol


ACCOUNTANT_REASON = "Đồng bộ từ file bảng công đã được Kế toán trưởng duyệt"


@dataclass
class WorkbookRow:
    sheet_name: str
    row_number: int
    full_name: str
    symbols: dict[date, str]
    payroll_work_days: float
    actual_work_days: float
    paid_leave_days: float
    unpaid_leave_days: float
    previous_leave_balance: float
    current_leave_credit: float
    remaining_leave_balance: float


def _key(value: object) -> str:
    text = unicodedata.normalize("NFD", str(value or ""))
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    return "".join(char for char in text.upper() if char.isalnum())


def _as_number(value: object, fallback: float = 0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return fallback


def _parse_period(sheet) -> tuple[date, date]:
    pattern = re.compile(r"(\d{2}/\d{2}/\d{4}).*?TO\s*(\d{2}/\d{2}/\d{4})", re.I)
    for row in sheet.iter_rows():
        for cell in row:
            match = pattern.search(str(cell.value or ""))
            if match:
                return (
                    date.fromisoformat("-".join(reversed(match.group(1).split("/")))),
                    date.fromisoformat("-".join(reversed(match.group(2).split("/")))),
                )
    raise ValueError("Không tìm thấy tiêu đề EMPLOYEE TIMESHEET có khoảng ngày.")


def _find_symbol_start_column(sheet, header_row: int, period_start: date, expected_days: int) -> int:
    candidates = [
        column
        for column in range(1, sheet.max_column + 1)
        if str(sheet.cell(header_row, column).value or "").strip().split("\n")[0] == str(period_start.day)
    ]
    if len(candidates) < 2:
        raise ValueError(f"{sheet.title}: không xác định được khối ký hiệu công thứ hai.")
    start = candidates[-1]
    if start + expected_days - 1 > sheet.max_column:
        raise ValueError(f"{sheet.title}: khối ký hiệu công không đủ số ngày của chu kỳ.")
    return start


def _find_column(sheet, header_row: int, header_key: str) -> int:
    for column in range(1, sheet.max_column + 1):
        if _key(sheet.cell(header_row, column).value) == header_key:
            return column
    raise ValueError(f"{sheet.title}: không tìm thấy cột {header_key!r}.")


def _read_rows(workbook, sheet_names: list[str]) -> tuple[date, date, list[WorkbookRow]]:
    if "Timesheet" not in workbook.sheetnames:
        raise ValueError("Không tìm thấy sheet 'Timesheet' để xác định kỳ công.")
    period_start, period_end = _parse_period(workbook["Timesheet"])
    period_dates = [
        date.fromordinal(day)
        for day in range(period_start.toordinal(), period_end.toordinal() + 1)
    ]
    result: list[WorkbookRow] = []
    seen_names: set[str] = set()

    for sheet_name in sheet_names:
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        header_row = 7 if sheet_name == "Timesheet" else 1
        symbol_start = _find_symbol_start_column(sheet, header_row, period_start, len(period_dates))
        columns = {
            "payroll": _find_column(sheet, header_row, "NGAYCONG"),
            "actual": _find_column(sheet, header_row, "NGAYCONGTT"),
            "unpaid": _find_column(sheet, header_row, "NGHIKHONGLUONG"),
            "paid": _find_column(sheet, header_row, "NGHIHUONGLUONG"),
            "previous": _find_column(sheet, header_row, "NGAYPHEPCONLAITHANGTRUOC"),
            "current": _find_column(sheet, header_row, "NGAYPHEPTHANGNAY"),
            "remaining": _find_column(sheet, header_row, "NGAYPHEPCONLAI"),
        }

        for row_number in range(header_row + 1, sheet.max_row + 1):
            full_name = str(sheet.cell(row_number, 2).value or "").strip()
            if not full_name:
                continue
            name_key = _key(full_name)
            if name_key in seen_names:
                raise ValueError(f"Nhân viên {full_name!r} xuất hiện nhiều lần trong workbook.")
            seen_names.add(name_key)
            result.append(
                WorkbookRow(
                    sheet_name=sheet_name,
                    row_number=row_number,
                    full_name=full_name,
                    symbols={
                        work_date: _clean_symbol(sheet.cell(row_number, symbol_start + offset).value)
                        for offset, work_date in enumerate(period_dates)
                    },
                    payroll_work_days=_as_number(sheet.cell(row_number, columns["payroll"]).value),
                    actual_work_days=_as_number(sheet.cell(row_number, columns["actual"]).value),
                    unpaid_leave_days=_as_number(sheet.cell(row_number, columns["unpaid"]).value),
                    paid_leave_days=_as_number(sheet.cell(row_number, columns["paid"]).value),
                    previous_leave_balance=_as_number(sheet.cell(row_number, columns["previous"]).value),
                    current_leave_credit=_as_number(sheet.cell(row_number, columns["current"]).value),
                    remaining_leave_balance=_as_number(sheet.cell(row_number, columns["remaining"]).value),
                )
            )
    return period_start, period_end, result


def import_final_template(
    path: Path,
    *,
    sheet_names: list[str] | None = None,
    working_days: dict[date, str] | None = None,
    employee_names: list[str] | None = None,
    payroll_days_only: bool = False,
) -> dict:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        period_start, period_end, source_rows = _read_rows(
            workbook,
            sheet_names or ["Timesheet", "TTS"],
        )
    finally:
        workbook.close()

    requested_employee_keys = {_key(name) for name in (employee_names or []) if name.strip()}
    selective_import = bool(requested_employee_keys)
    if selective_import:
        source_rows = [source for source in source_rows if _key(source.full_name) in requested_employee_keys]
        found_employee_keys = {_key(source.full_name) for source in source_rows}
        missing_employee_keys = requested_employee_keys - found_employee_keys
        if missing_employee_keys:
            raise ValueError(
                "Không tìm thấy trong file: "
                + ", ".join(sorted(name for name in (employee_names or []) if _key(name) in missing_employee_keys))
            )

    working_days = working_days or {}
    for work_date in working_days:
        if not (period_start <= work_date <= period_end):
            raise ValueError(f"Ngày làm bù {work_date.isoformat()} nằm ngoài kỳ công.")

    db = SessionLocal()
    imported = 0
    unmatched: list[dict[str, object]] = []
    excluded: list[str] = []
    salary_period = period_end.strftime("%Y-%m")
    try:
        for work_date, reason in (() if payroll_days_only else working_days.items()):
            calendar_row = db.query(HolidaySetting).filter(HolidaySetting.holiday_date == work_date).first()
            if calendar_row is None:
                calendar_row = HolidaySetting(
                    holiday_name=reason,
                    holiday_date=work_date,
                    is_custom=True,
                    is_working_day=True,
                )
                db.add(calendar_row)
            else:
                calendar_row.holiday_name = reason
                calendar_row.is_custom = True
                calendar_row.is_working_day = True

        employees = db.query(Employee).all()
        employees_by_name = {_key(employee.full_name): employee for employee in employees}
        employees_by_id = {employee.id: employee for employee in employees}
        imported_employee_ids: set[int] = set()

        for source in source_rows:
            employee = employees_by_name.get(_key(source.full_name))
            if employee is None:
                unmatched.append({"name": source.full_name, "sheet": source.sheet_name, "row": source.row_number})
                continue

            imported_employee_ids.add(employee.id)
            timesheet = (
                db.query(Timesheet)
                .filter(
                    Timesheet.employee_id == employee.id,
                    Timesheet.period_start == period_start,
                    Timesheet.period_end == period_end,
                )
                .first()
            )
            if payroll_days_only:
                if timesheet is None:
                    unmatched.append(
                        {
                            "name": source.full_name,
                            "sheet": source.sheet_name,
                            "row": source.row_number,
                            "reason": "Không có bảng công hiện hữu trong kỳ",
                        }
                    )
                    continue
                timesheet.total_payroll_days = source.payroll_work_days
                salary_input = (
                    db.query(MonthlySalaryInput)
                    .filter(
                        MonthlySalaryInput.employee_id == employee.id,
                        MonthlySalaryInput.salary_period == salary_period,
                    )
                    .first()
                )
                if salary_input is None:
                    db.add(
                        MonthlySalaryInput(
                            employee_id=employee.id,
                            salary_period=salary_period,
                            actual_working_days=source.payroll_work_days,
                        )
                    )
                else:
                    salary_input.actual_working_days = source.payroll_work_days
                imported += 1
                continue

            if timesheet is None:
                timesheet = Timesheet(employee_id=employee.id, period_start=period_start, period_end=period_end)
                db.add(timesheet)
                db.flush()
            timesheet.approval_status = "approved"

            business_total = 0.0
            for work_date, workbook_symbol in source.symbols.items():
                final_symbol = "" if work_date.weekday() >= 5 and work_date not in working_days else workbook_symbol
                daily = (
                    db.query(AttendanceDaily)
                    .filter(AttendanceDaily.employee_id == employee.id, AttendanceDaily.work_date == work_date)
                    .first()
                )
                old_symbol = _clean_symbol(daily.attendance_symbol) if daily else ""
                if daily is None:
                    daily = AttendanceDaily(
                        employee_id=employee.id,
                        work_date=work_date,
                        period_start=period_start,
                        period_end=period_end,
                        attendance_symbol=final_symbol,
                        abnormal_level=None,
                        source_priority=3,
                    )
                    db.add(daily)
                else:
                    daily.period_start = period_start
                    daily.period_end = period_end
                    daily.attendance_symbol = final_symbol
                    daily.source_priority = 3

                entry = (
                    db.query(TimesheetEntry)
                    .filter(
                        TimesheetEntry.timesheet_id == timesheet.id,
                        TimesheetEntry.employee_id == employee.id,
                        TimesheetEntry.work_date == work_date,
                    )
                    .first()
                )
                override_reason = ACCOUNTANT_REASON
                if work_date in working_days:
                    override_reason = f"{ACCOUNTANT_REASON} | {working_days[work_date]}"
                if entry is None:
                    entry = TimesheetEntry(
                        timesheet_id=timesheet.id,
                        employee_id=employee.id,
                        work_date=work_date,
                        original_symbol=old_symbol,
                        final_symbol=final_symbol,
                        is_overridden=True,
                        override_reason=override_reason,
                        overridden_at=datetime.now(timezone.utc),
                    )
                    db.add(entry)
                else:
                    if not entry.original_symbol:
                        entry.original_symbol = old_symbol
                    entry.final_symbol = final_symbol
                    entry.is_overridden = True
                    entry.override_reason = override_reason
                    entry.overridden_at = datetime.now(timezone.utc)
                business_total += 1.0 if final_symbol == "CT" else 0.0

            # Preserve the accountant-approved hardcoded totals and leave
            # snapshots exactly, even if they do not equal a formula recalc.
            timesheet.total_payroll_days = source.payroll_work_days
            timesheet.total_work_days = source.actual_work_days
            timesheet.total_paid_leave_days = source.paid_leave_days
            timesheet.total_unpaid_leave_days = source.unpaid_leave_days
            timesheet.total_absent_days = source.unpaid_leave_days
            timesheet.total_business_trip_days = business_total
            timesheet.previous_paid_leave_balance = source.previous_leave_balance
            timesheet.current_month_paid_leave_credit = source.current_leave_credit
            timesheet.remaining_paid_leave_days = source.remaining_leave_balance

            employee.paid_leave_balance = source.remaining_leave_balance
            if str(employee.employee_type or "FULLTIME").upper() == "FULLTIME":
                employee.annual_leave_quota = source.current_leave_credit * 12

            salary_input = (
                db.query(MonthlySalaryInput)
                .filter(
                    MonthlySalaryInput.employee_id == employee.id,
                    MonthlySalaryInput.salary_period == salary_period,
                )
                .first()
            )
            payable_days = source.payroll_work_days
            if salary_input is None:
                salary_input = MonthlySalaryInput(
                    employee_id=employee.id,
                    salary_period=salary_period,
                    actual_working_days=payable_days,
                )
                db.add(salary_input)
            else:
                salary_input.actual_working_days = payable_days
            imported += 1

        if not selective_import and not payroll_days_only:
            current_timesheets = (
                db.query(Timesheet)
                .filter(Timesheet.period_start == period_start, Timesheet.period_end == period_end)
                .all()
            )
            for timesheet in current_timesheets:
                if timesheet.employee_id in imported_employee_ids:
                    continue
                timesheet.approval_status = "excluded"
                employee = employees_by_id.get(timesheet.employee_id)
                if employee is not None:
                    excluded.append(employee.full_name)
                salary_input = (
                    db.query(MonthlySalaryInput)
                    .filter(
                        MonthlySalaryInput.employee_id == timesheet.employee_id,
                        MonthlySalaryInput.salary_period == salary_period,
                    )
                    .first()
                )
                if salary_input is not None:
                    salary_input.actual_working_days = 0.0

            # Salary rows may already exist even when no timesheet was generated
            # for that employee. Keep the rows for audit/history, but do not let a
            # stale default day count leak into the accountant's August payroll.
            stale_salary_query = db.query(MonthlySalaryInput).filter(
                MonthlySalaryInput.salary_period == salary_period
            )
            if imported_employee_ids:
                stale_salary_query = stale_salary_query.filter(
                    ~MonthlySalaryInput.employee_id.in_(sorted(imported_employee_ids))
                )
            stale_salary_rows = stale_salary_query.all()
            for salary_input in stale_salary_rows:
                salary_input.actual_working_days = 0.0

        if not payroll_days_only:
            period = (
                db.query(TimesheetPeriod)
                .filter(TimesheetPeriod.period_start == period_start, TimesheetPeriod.period_end == period_end)
                .first()
            )
            if period is None:
                period = TimesheetPeriod(period_start=period_start, period_end=period_end)
                db.add(period)
            period.is_locked = True
            period.locked_by_user_id = None
            period.locked_at = datetime.now(timezone.utc)

        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()

    return {
        "period": f"{period_start.isoformat()} to {period_end.isoformat()}",
        "sheets": sheet_names or ["Timesheet", "TTS"],
        "source_rows": len(source_rows),
        "selective_import": selective_import,
        "payroll_days_only": payroll_days_only,
        "selected_employees": employee_names or [],
        "imported": imported,
        "unmatched": unmatched,
        "excluded": sorted(excluded),
        "working_days": [
            {"date": work_date.isoformat(), "reason": reason}
            for work_date, reason in sorted(working_days.items())
        ],
        "salary_period": salary_period,
    }


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Đồng bộ bảng công đã được Kế toán trưởng duyệt")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--sheets", default="Timesheet,TTS")
    parser.add_argument("--working-day", action="append", default=[], help="Ngày cuối tuần đi làm, YYYY-MM-DD")
    parser.add_argument("--working-day-reason", default="Làm bù")
    parser.add_argument("--employee", action="append", default=[], help="Chỉ đồng bộ đúng nhân viên này")
    parser.add_argument(
        "--payroll-days-only",
        action="store_true",
        help="Chỉ lưu cột Ngày công và đồng bộ số ngày tính lương; không sửa ký hiệu/nghỉ phép/trạng thái",
    )
    args = parser.parse_args()
    result = import_final_template(
        args.workbook,
        sheet_names=[item.strip() for item in args.sheets.split(",") if item.strip()],
        working_days={date.fromisoformat(value): args.working_day_reason for value in args.working_day},
        employee_names=args.employee,
        payroll_days_only=args.payroll_days_only,
    )
    print(json.dumps(result, ensure_ascii=True, default=str))
