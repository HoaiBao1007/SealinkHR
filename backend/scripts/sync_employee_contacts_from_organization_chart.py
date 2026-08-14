"""Synchronize employee contact details from the approved organization workbook.

The workbook uses English/Notion display names while the database can contain
legacy spellings.  Matching is therefore pinned to the biometric machine ID for
each reviewed row instead of relying on fuzzy name matching.

Run a dry-run first:
    python scripts/sync_employee_contacts_from_organization_chart.py

Apply the reviewed changes:
    python scripts/sync_employee_contacts_from_organization_chart.py --apply
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


BACKEND_DIR = Path(__file__).resolve().parents[1]
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from app.db.session import SessionLocal  # noqa: E402
from app.models.employee import Employee  # noqa: E402


DEFAULT_WORKBOOK = Path(
    r"C:\Users\hoaib\Downloads\Sea_Link_Organize_Chart_2026-07-01.xlsx"
)
SOURCE_SHEET = "Danh sách nhân sự"

# Reviewed against the employee database on 2026-07-30.  These stable machine
# IDs also cover spelling differences such as QINNIE/QUINNIE and BRIAN/BRAIN.
MACHINE_ID_BY_SOURCE_NAME = {
    "IVY VAN": "12",
    "KIEN TON": "15",
    "MICHAEL PHUONG": "68",
    "RUBY HUONG": "29",
    "YAYAA NGAN": "50",
    "SABER NGAN": "40",
    "ERIC QUAN": "64",
    "VICKY QUYEN": "8",
    "WILLIAM TRUNG": "2",
    "LISA NGA": "4",
    "DINO LONG": "21",
    "JACOB QUANG": "28",
    "BOO BAO": "20",
    "HARRY TUE": "66",
    "SUN NHAT": "63",
    "MARVIN THIEN": "18",
    "JOYCE LANG": "17",
    "DUKE PHONG": "13",
    "PARADO QUANG": "36",
    "QINNIE QUYEN": "6",
    "EDWARD DUONG": "70",
    "JUDY THAO": "59",
    "HILMAR HOA": "41",
    "LUCIUS TUONG": "23",
    "JOLIE BICH": "7",
    "KATHIE TU": "3",
    "FIONA PHUONG": "65",
    "SANDRA NGA": "56",
    "LAYLA HA": "27",
    "CATHAY NGOC": "31",
    "NELLY NHU": "5",
    "SAM VAN": "16",
    "ASHLEY NHI": "51",
    "AHRI VAN": "74",
    "SARAH DUNG": "44",
    "DANIEL BAO": "39",
    "AXEL HOA": "60",
    "LINN LINH": "54",
    "JAYCE TAM": "69",
    "BRIAN TUONG": "14",
    "CHEESE CHI": "75",
    "DAISY NHI": "22",
    "LUCY THAO": "35",
    "LILY LINH": "72",
    "YUYU LINH": "30",
    "TOMMY DAT": "26",
    "LUNA THAO": "37",
    "BARON BAO": "38",
}

# The source workbook itself flags this typo for verification.  Keep the
# correction explicit and auditable instead of silently applying heuristics.
EMAIL_CORRECTIONS = {
    "nguyen.bao@sea-lik.com": "nguyen.bao@sea-link.com",
}


@dataclass(frozen=True)
class SourceContact:
    source_name: str
    email: str | None
    personal_phone: str | None
    company_phone: str | None


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _normalize_email(value: object) -> str | None:
    email = _text(value).lower()
    if not email:
        return None
    return EMAIL_CORRECTIONS.get(email, email)


def _normalize_phone(value: object) -> str | None:
    digits = re.sub(r"\D", "", _text(value))
    if not digits:
        return None
    # Excel numeric cells drop the leading zero from Vietnamese mobile numbers.
    if len(digits) == 9:
        digits = f"0{digits}"
    if len(digits) == 10:
        return f"{digits[:4]} {digits[4:7]} {digits[7:]}"
    if len(digits) == 11:
        return f"{digits[:4]} {digits[4:7]} {digits[7:]}"
    return digits


def _excel_phone(cell: object) -> str | None:
    value = getattr(cell, "value", None)
    number_format = str(getattr(cell, "number_format", "") or "")
    # Preserve the leading zero encoded by Excel's custom numeric format.
    if isinstance(value, (int, float)) and re.fullmatch(r"0+", number_format):
        value = f"{int(value):0{len(number_format)}d}"
    return _normalize_phone(value)


def load_contacts(workbook_path: Path) -> list[SourceContact]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[SOURCE_SHEET]
        contacts: list[SourceContact] = []
        for row in worksheet.iter_rows(min_row=5):
            source_name = _text(row[2].value).upper()
            if not source_name:
                continue
            contacts.append(
                SourceContact(
                    source_name=source_name,
                    email=_normalize_email(row[4].value),
                    personal_phone=_excel_phone(row[5]),
                    company_phone=_excel_phone(row[6]),
                )
            )
        return contacts
    finally:
        workbook.close()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Commit changes. Without this flag the command is read-only.",
    )
    args = parser.parse_args()

    if not args.workbook.exists():
        raise SystemExit(f"Không tìm thấy file nguồn: {args.workbook}")

    contacts = load_contacts(args.workbook)
    source_names = {contact.source_name for contact in contacts}
    configured_names = set(MACHINE_ID_BY_SOURCE_NAME)
    missing_mapping = sorted(source_names - configured_names)
    stale_mapping = sorted(configured_names - source_names)
    if missing_mapping or stale_mapping:
        raise SystemExit(
            "Mapping không còn khớp file nguồn. "
            f"Thiếu mapping={missing_mapping}; mapping dư={stale_mapping}"
        )

    with SessionLocal() as db:
        employees = {
            employee.machine_employee_id: employee
            for employee in db.query(Employee)
            .filter(
                Employee.machine_employee_id.in_(
                    list(MACHINE_ID_BY_SOURCE_NAME.values())
                )
            )
            .all()
        }
        missing_employees: list[str] = []
        changes: list[str] = []
        unchanged = 0

        for contact in contacts:
            machine_id = MACHINE_ID_BY_SOURCE_NAME[contact.source_name]
            employee = employees.get(machine_id)
            if employee is None:
                missing_employees.append(
                    f"{contact.source_name} (mã máy {machine_id})"
                )
                continue

            before = (
                employee.company_email,
                employee.phone_number,
                employee.company_phone_number,
            )
            after = (
                contact.email,
                contact.personal_phone,
                contact.company_phone,
            )
            if before == after:
                unchanged += 1
                continue

            changes.append(
                f"- {employee.full_name} [{machine_id}] / {contact.source_name}: "
                f"email {before[0]!r} -> {after[0]!r}; "
                f"SĐT cá nhân {before[1]!r} -> {after[1]!r}; "
                f"SĐT công ty {before[2]!r} -> {after[2]!r}"
            )
            employee.company_email = contact.email
            employee.phone_number = contact.personal_phone
            employee.company_phone_number = contact.company_phone

        if missing_employees:
            db.rollback()
            raise SystemExit(
                "Không cập nhật vì thiếu hồ sơ trong database: "
                + ", ".join(missing_employees)
            )

        print(
            f"Đã đối chiếu {len(contacts)} nhân viên; "
            f"{len(changes)} hồ sơ cần cập nhật; {unchanged} hồ sơ không đổi."
        )
        for change in changes:
            print(change)

        if args.apply:
            db.commit()
            print("ĐÃ COMMIT dữ liệu liên hệ vào database.")
        else:
            db.rollback()
            print("DRY-RUN: chưa thay đổi database. Dùng --apply để xác nhận.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
